#!/usr/bin/env python3
# -*- coding: UTF-8 -*-
"""
Description: A helper/wrapper library to aid in using openpyxl as a data structure.

Usage: See below. Like at the bottom.

License: See main program.

##stop reading now##

"""
#set defaults
printStuff=True
debug=False
#debug=True
consoleEncoding='utf-8'

#These must be here or the library will crash even if these modules have already been imported by main program.
import os.path                            #Extract extension from filename, and test if file exists.
from pathlib import Path            #Override file in file system with another and create subfolders.
import sys                                   #End program on fail condition.
import openpyxl                          #Used as the core internal data structure and also to read/write xlsx files.
import codecs                            #Improves error handling when dealing with text file codecs.
import csv                                    #Read and write to csv files. Example: Read in 'resources/languageCodes.csv'
try:
    import odfpy                           #Provides interoperability for Open Document Spreadsheet (.ods).
    odfpyLibraryIsAvailable=True
except:
    odfpyLibraryIsAvailable=False
try:
    import xlrd                              #Provides reading from Microsoft Excel Document (.xls).
    xlrdLibraryIsAvailable=True
except:
    xlrdLibraryIsAvailable=False
try:
    import xlwt                              #Provides writing to Microsoft Excel Document (.xls).
    xlwtLibraryIsAvailable=True
except:
    xlwtLibraryIsAvailable=False


#wrapper class for spreadsheet data structure
class Strawberry:
    def __init__(self, myFileName=None, myFileNameEncoding=None, ignoreWhitespaceForCSV=False): #self is not a keyword. It can be anything, like pie, but it must be the first argument for every function in the class. Quirk: It can be different string/word for each method and they all still refer to the same object.
        self.workbook = openpyxl.Workbook()
        self.spreadsheet = self.workbook.active
        if myFileName != None:
            if myFileNameEncoding == None:
                sys.exit( ('Please specify an encoding for: '+myFileName).encode(consoleEncoding) )
            #Then find the extension of the file.
            myFileNameOnly, myFileExtensionOnly = os.path.splitext(myFileName)
            #If there is no extension, then crash.
            if myFileExtensionOnly == '':
                sys.exit(   ('Warning: Cannot instantiate class using a file that lacks an extension. Reference:\''+myFileName+'\'').encode(consoleEncoding)   )
            #Check to make sure file exists.
            #If file does not exist, then crash.
            if os.path.isfile(myFileName) != True:
                sys.exit(('Warning: This file does not exist:\''+myFileName+'\'').encode(consoleEncoding))
            #If extension = .csv, then call importFromCSV(myFileName)
            if myFileExtensionOnly == '.csv':
                self.importFromCSV(myFileName, myFileNameEncoding, ignoreWhitespaceForCSV)
            #if extension = .xlsx, then call importFromXLSX(myFileName)
            elif myFileExtensionOnly == '.xlsx':
                self.importFromXLSX(myFileName, myFileNameEncoding)
            elif myFileExtensionOnly == '.xls':
                self.importFromXLS(myFileName, myFileNameEncoding)
            elif myFileExtensionOnly == '.ods':
                self.importFromODS(myFileName, myFileNameEncoding)
            else:
                sys.exit(('Warning: Cannot instantiate class using file with unknown extension:'+myFileExtensionOnly+' Reference:\''+myFileName+'\'').encode(consoleEncoding))

    def __str__(self):
        #maybe return the headers from the spreadsheet?
        #return str(spreadsheet[1])
        #return 'pie'
        return str(getRow(1))

    #expects a python list
    def appendRow(self,newRow):
        self.spreadsheet.append(newRow)

    #def appendColumn(self, newColumn) #Does not seem to be needed. Data is just not processed that way.

    #Full name of this function is getCellAddressFromRawCellString, but was shortened for legibility. Edit: Made it longer again.
    #This functions would return 'B5' from: <Cell 'Sheet'.B5>
    def getCellAddressFromRawCellString(self, myInputCellRaw):
        #print('raw cell data='+str(myInputCellRaw))
        #myInputCellRaw=str(myInputCellRaw)
        #Basically, split the string according to . and then split it again according to > to get back only the CellAddress
        return str(myInputCellRaw).split('.', maxsplit=1)[1].split('>')[0]
        #return [currentRow, currentColumn

    #This function returns a list containing 2 strings that represent a row and column extracted from input Cell address
    #such as returning ['5', 'B'] from: <Cell 'Sheet'.B5>   It also works for complicated cases like AB534.
    def getRowAndColumnFromRawCellString(self, myInputCellRaw):
        #print('raw cell data='+str(myInputCellRaw))
        #basically, split the string according to . and then split it again according to > to get back only the CellAddress
        #myInputCell=str(myInputCellRaw).split('.', maxsplit=1)[1].split('>')[0]
        myInputCell=self.getCellAddressFromRawCellString(myInputCellRaw)
        index=0
        for i in range(10):
            try:
                int(myInputCell[index:index+1])
                index=i
                break #this break and the assignment above will only execute when the int conversion works
            except:
                #this will execute if there is an error, like int('A')
                #this will not execute if the int conversion succeds
                #print('index='+str(index))
                pass
            index+=1
        currentColumn=myInputCell[:index]#does not include character in string[Index] because index is after the :
        currentRow=myInputCell[index:]#includes character specified by string[index] because index is before the :
        #return ['pie', 'apple']
        return [currentRow, currentColumn]

    #Example:
    #myRawCell=''
    #for row in mySpreadsheet:
    #    for i in row:
    #        if i.value == 'lots of pies':
    #            print(str(i) + '=' + str(i.value))
    #            myRawCell=i
    #currentRow, currentColumn = spreadsheet.getRowAndColumnFromRawCellString(myRawCell)


    #Returns a list with the contents of the row number specified.
    #Should return None for any blank entry as in: ['pie', None, 'lots of pies']
    def getRow(self, rowNumber):
        #print(rowNumber)
        #return spreadsheet[rowNumber] #returns the raw cell addresses instead of the values.
        #returns the values in a list
        myList=[]
        for cell in self.spreadsheet[rowNumber]:
            if debug == True:
                print( (str(self.spreadsheet[self.getCellAddressFromRawCellString(cell)].value)+',').encode(consoleEncoding),end='')
            myList.append(self.spreadsheet[self.getCellAddressFromRawCellString(cell)].value)
        if debug == True:
            print('')
        return myList


    #Returns a list with the contents of the column specified (by letter). 
    #Should return None for any blank entry as in: ['pie', None, 'lots of pies']
    def getColumn(self, columnLetter):
        myList=[]
        for cell in self.spreadsheet[columnLetter]:
            #print(str(mySpreadsheet[self.getCellAddressFromRawCellString(cell)].value)+',',end='')
            #myList[i]=mySpreadsheet[self.getCellAddressFromRawCellString(cell)].value  #Doesn't work due to out of index error. Use append() method.
            myList.append(self.spreadsheet[self.getCellAddressFromRawCellString(cell)].value)
        return myList
        #print("Hello, world!")

    #case sensitive
    #def getColumnLetterFromSearchString():
        #No. Just search normally, and search should always return a list with the column and row seperately.

    #helper function that changes the data for a row in mySpreadsheet to what is specified in a python List []
    #Note: This is only for modifying existing rows. To add a brand new row, use append:
        #Example: newRow = ['pies', 'lots of pies']
        #mySpreadsheet.append(newRow)
    #The rowLocation specified is the nth rowLocation, not the [0,1,2,3...row] because rows start with 1
    def replaceRow(self, newRowList, rowLocation):
        if debug == True:
            print(str(len(newRowList)).encode(consoleEncoding))
            print(str(range(len(newRowList))).encode(consoleEncoding))
        for i in range(len(newRowList)):
            #Syntax for assignment is: mySpreadsheet['A4'] = 'pie'
            #mySpreadsheet['A4'] without an assignment returns: <Cell 'Sheet'.A4> 
            #columns begin with 1 instead of 0, so add 1 when referencing the target column, but not the source because source is a python list which are referenced as list[0], list[1], list[2], list[3], etc
            #Was workaround for Syntax error cannot assign value to function call: mySpreadsheet.cell(row=5, column=3)='pies'  
            #spreadsheet[getCellAddressFromRawCellString(spreadsheet.cell(row=int(rowLocation), column=i+1))]=newRowList[i]
            #A more direct way of doing the same thing is to use .value without () on the cell after the cell reference.
            self.spreadsheet.cell(row=int(rowLocation), column=i+1).value=newRowList[i]
        #return myWorkbook

    #Example: replaceRow(newRow,7)


    def replaceColumn(self, newColumnInAList, columnLetter):
        #So, how to convert a columnLetter into a number or does column='A' also work?
        #Answer column='A' does not work but there are some built in methods:
        #x = openpyxl.utils.column_index_from_string('A')   #returns 1 as an int
        #y= openpyxl.utils.get_column_letter(1)   #returns 'A'
        #Example: mySpreadsheet.cell(row=3, column=openpyxl.utils.column_index_from_string('B')).value='pies'
        #Documentation: https://openpyxl.readthedocs.io/en/stable/api/openpyxl.utils.cell.html
        if debug == True:
            print(( 'Replacing column \''+columnLetter+'\' with the following contents:').encode(consoleEncoding))
            print(str(newColumnInAList).encode(consoleEncoding))
        for i in range(len(newRowList)):
            #Syntax for assignment is: mySpreadsheet['A4'] = 'pie''
            #Rows begin with 1, not 0, so add 1 to the reference row, but not to source list since list starts references at 0.
            self.spreadsheet.cell(row=int(i+1), column=openpyxl.utils.column_index_from_string(columnLetter)).value=newRowList[i]

    #Example: replaceColumn(newColumn,7)


    #Return either None if there is no cell with the search term, or the column letter of the cell if it found it. Case and whitespace sensitive search.
    #Aside: To determine the row, the column, or both from the raw cell address, call self.getRowAndColumnFromRawCellString(rawCellAddress)
    def searchHeaders(self, searchTerm):
        cellFound=None
        for row in self.spreadsheet[1]:
            for i in row:
                if i.value == searchTerm:
                    cellFound=i
            #if cellFound != None:
            #    print('found')
            #else:
            #    print('notfound')
            break #stop searching after first row
        if cellFound == None:
            return None
        #Slower.
        #else:
            #myRowNumber, myColumnLetter = self.getRowAndColumnFromRawCellString(cellFound)
        #return myColumnLetter
        return self.getRowAndColumnFromRawCellString(cellFound)[1]   #Faster.

    #Example:
    #cellFound=None
    #isFound=searchHeader(mySpreadsheet,searchTerm)
    #if isFound == None:
    #    print('was not found')
    #else:
    #    print('searchTerm:\"'+searchTerm+'" was found at:'+str(isFound))


    #This returns either [None, None] if there is no cell with the search term, or a list containing the [row, column] (the address). Case and whitespace sensitive.
    #To determine the row, the column, or both from the raw cell address, use self.getRowAndColumnFromRawCellString(rawCellAddress)
    def searchSpreadsheet(self, searchTerm):
        for row in self.spreadsheet.iter_rows():
            for cell in row:
                if cell.value == searchTerm:
                    return self.getRowAndColumnFromRawCellString(cell)
        return [None, None]


    #These return either [None,None] if there is no cell with the search term, or a [list] containing the cell row and the cell column (the address in a list). Case insensitive. Whitespace sensitive.
    #To determine the row, the column, or both from the raw cell address, use self.getRowAndColumnFromRawCellString(rawCellAddress)
    def searchRowsCaseInsensitive(self, searchTerm):
        for row in self.spreadsheet.iter_rows():
            for cell in row:
                if isinstance(cell.value, str):
                    if cell.value.lower() == str(searchTerm).lower():
                        return self.getRowAndColumnFromRawCellString(cell)
        return [None, None]

    def searchColumnsCaseInsensitive(self, searchTerm):
        for column in self.spreadsheet.iter_cols():
            for cell in column:
                if isinstance(cell.value, str):
                    if cell.value.lower() == str(searchTerm).lower():
                        return self.getRowAndColumnFromRawCellString(cell)
        return [None, None]


    #Give this function a spreadsheet object (subclass of workbook) and it will print the contents of that sheet
    def printAllTheThings(self):
        for row in self.spreadsheet.iter_rows(min_row=1, values_only=True):
            temp=''
            for cell in row:
                temp=temp+','+str(cell)
            print(temp[1:].encode(consoleEncoding))#ignore first , in output

    #Old example: printAllTheThings(mySpreadsheet)
    #New syntax: 
    #mySpreadsheet= Strawberry()
    #mySpreadsheet.printAllTheThings()


    #TODO:
    #1) Export an existing spreadsheet to a file.
    #2) Import a file into an existing spreadsheet.
    #References/objects are done using workbooks, not the active spreadsheet.
    #Edit: Return value/reference for reading from files should be done by returning a class instance (object) of Strawberry()
    #Strawberry should have its own methods for writing to files of various formats.
    #All files follow the same rule of the first row being reserved for header values and invalid for inputting/outputting actual data.

    def importFromCSV(self, fileNameWithPath,myFileNameEncoding,ignoreWhitespace=True):
        #import languageCodes.csv, but first check to see if it exists
        if os.path.isfile(fileNameWithPath) != True:
            sys.exit(('\n Error. Unable to find .csv file:"' + fileNameWithPath + '"' + usageHelp).encode(consoleEncoding))

        #tempWorkbook = openpyxl.Workbook()
        #tempSpreadsheet = tempWorkbook.active
        #tempSpreadsheet = Strawberry()

        #It looks like quoting fields in csv's that use commas , and new
        #lines works but only as double quotes " and not single quotes '
        #Spaces are also preserved as-is if they are within the commas (,) by default, so remove them
        #If spaces are intended to be within the entry, then the user can encapslate them in double quotes
        #Need to test. Even double quotes might not preserve them. Tested: They do not.
        #Could also just say not supported since it is almost certainly an error for hand-written CSV's.
        #Could also have a flag that switches back and forth.
        #Partial solution, added "ignoreWhitespace" function parameter which defaults to True.
        #Reading from dictionaries can be called with the "False" option for maximum flexibility.
        #New problem: How to expose this functionality to user?
        with open(fileNameWithPath, newline='', encoding=myFileNameEncoding) as myFile:
            csvReader = csv.reader(myFile)
            for line in csvReader:
                if debug == True:
                    print(str(line).encode(consoleEncoding))
                #clean up whitespace for entities
                if ignoreWhitespace == True:
                    for i in range(len(line)):
                        line[i]=line[i].strip()
                #tempSpreadsheet.append(line)
                #tempSpreadsheet.appendRow(line)
                self.spreadsheet.append(line)
        #return tempWorkbook
        if debug == True:
            self.printAllTheThings()

    def exportToCSV(self, fileNameWithPath, myFileNameEncoding):
        print('Hello World'.encode(consoleEncoding))


    def importFromXLSX(self, fileNameWithPath, myFileNameEncoding):
        print('Hello World'.encode(consoleEncoding))
        #return workbook
    def exportToXLSX(self, fileNameWithPath, myFileNameEncoding):
        print('Hello World'.encode(consoleEncoding))
        #theWorkbook.save(filename="myAwesomeSpreadsheet.xlsx")
        workBook.save(filename=fileNameWithPath)


    def importFromXLS(self, fileNameWithPath, myFileNameEncoding):
        print('Hello World'.encode(consoleEncoding))
        #return workbook
    def exportToXLS(self, fileNameWithPath, myFileNameEncoding):
        print('Hello World'.encode(consoleEncoding))


    def importFromODS(self, fileNameWithPath, myFileNameEncoding):
        print('Hello World'.encode(consoleEncoding))
        #return workbook
    def exportToODS(self, fileNameWithPath, myFileNameEncoding):
        print('Hello World'.encode(consoleEncoding))













"""
#Usage examples, assuming this library is in a subfolder named 'resources':
defaultEncoding='utf-8'
myFileName = 'myFile.txt'

import openpyxlHelpers

spreadsheet=openpyxlHelpers.Strawberry()

searchCellRow, searchCellColumn = spreadsheet.search


if dealWithEncodingLibraryIsAvailable == True:
    #Update internal library variables to match main program settings.
    dealWithEncoding.debug=debug
    dealWithEncoding.consoleEncoding=consoleEncoding

    inputEncodingType = dealWithEncoding.ofThisFile(myFileName=inputFileName, rawCommandLineOption=command_Line_arguments.inputEncoding, fallbackEncoding=defaultEncodingType)

    #or, to use only positional arguments
    inputEncodingType = dealWithEncoding.ofThisFile(inputFileName, command_Line_arguments.inputEncoding, defaultEncodingType)

    #or, To detect encoding of a file with the chardet library that has already been determined to exist, and does not consider user preferences, fallback encoding, or a return of None from the chardet library:
    inputEncodingType= dealWithEncoding.detectEncoding(inputFileName)
else:
    inputEncodingType=defaultEncoding

print(inputFileName+' will use encoding type: '+inputEncodingType)

"""
