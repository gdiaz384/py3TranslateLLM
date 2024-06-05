#!/usr/bin/env python3
# -*- coding: UTF-8 -*-
"""
Description: A helper/wrapper library to aid in using openpyxl as a data structure. Supports i/o for .csv, .xlsx, .xlsx, .ods. If using this library for cache.xlsx, where entries are unique, there are specialized functions available as well.

Usage: See below. Like at the bottom.

Copyright (c) 2024 gdiaz384; License: See main program.

"""
__version__='2024.06.04'

#set defaults
#printStuff=True
verbose=False
debug=False
#debug=True
consoleEncoding='utf-8'
defaultTextFileEncoding='utf-8'   # Settings that should not be left as a default setting should have default prepended to them.
inputErrorHandling='strict'
#outputErrorHandling='namereplace'  #This is set dynamically below.

#These must be here or the library will crash even if these modules have already been imported by main program.
import os.path                            # Extract extension from filename, and test if file exists.
import pathlib                             # For pathlib.Path() Override file in file system with another and create subfolders.
import sys                                   # End program on fail condition.
#import random                             # Used to create random numbers. 
import openpyxl                          # Used as the core internal data structure and also to read/write xlsx files.
import csv                                   # Read and write to csv files. Example: Read in 'resources/languageCodes.csv'
try:
    import xlrd                              #Provides reading from Microsoft Excel Document (.xls).
    xlrdLibraryIsAvailable=True
except:
    xlrdLibraryIsAvailable=False
try:
    import xlwt                              #Provides writing to Microsoft Excel Document (.xls).
    xlwtLibraryIsAvailable=True
except:
    xlwtLibraryIsAvailable=False
try:
    import odfpy                           #Provides interoperability for Open Document Spreadsheet (.ods). Alternatives: https://github.com/renoyuan/easyofd pyexcel-ods3, pyexcel-ods, ezodf
    odfpyLibraryIsAvailable=True
except:
    odfpyLibraryIsAvailable=False

#Using the 'namereplace' error handler for text encoding requires Python 3.5+, so use an older one if necessary.
sysVersion=int(sys.version_info[1])
if sysVersion >= 5:
    outputErrorHandling='namereplace'
elif sysVersion < 5:
    outputErrorHandling='backslashreplace'    
else:
    sys.exit('Unspecified error.'.encode(consoleEncoding))


#wrapper class for spreadsheet data structure
class Strawberry:
    # self is not a keyword. It can be anything, like pie, but it must be the first argument for every function in the class. 
    # Quirk: It can be different string/word for each method and they all still refer to the same object.
    def __init__(self, myFileName=None, fileEncoding=defaultTextFileEncoding, removeWhitespaceForCSV=False, addHeaderToTextFile=False, spreadsheetNameInWorkbook=None, readOnlyMode=False, csvDialect=None):
        # https://openpyxl.readthedocs.io/en/stable/api/openpyxl.workbook.workbook.html
        self.fileEncoding=fileEncoding
        self.workbook = openpyxl.Workbook()
        if spreadsheetNameInWorkbook == None:
            self.spreadsheet = self.workbook.active
            self.spreadsheetName=self.spreadsheet.title
        else:
            self.spreadsheetName=spreadsheetNameInWorkbook
            #print(spreadsheetNameInWorkbook)
            self.workbook.create_sheet( title = self.spreadsheetName , index=0 )
            #print(self.workbook.sheetnames)
            self.spreadsheet = self.workbook[ self.spreadsheetName ]
        self.readOnlyMode = readOnlyMode
        self.csvDialect=csvDialect
        self.addHeaderToTextFile=addHeaderToTextFile
        #self.randomNumber=int( random.random() * 500000 )

        # These last two variables are only for use when chocolate.Strawberry() is being used as cache.xlsx. Ignore otherwise.
        # Index is every entry in the first column, A with an associated pointer, as an integer, to the correct row in the main spreadsheet.
        # Every item in the cache must be unique, not None, and not an empty string ''.
        self.index={}
        # the last entry i
        self.lastEntry=len(self.index)

        # Are there any use cases for creating a spreadsheet in memory without an associated file name? Since chocolate.Strawberry() is a data structure, this must be 'yes' by definition, but what is the use case for that exactly? When would it be useful to only create a spreadsheet in memory but never write it out?
        if myFileName != None:
            #if fileEncoding == None:
                #Actually, the encoding might be None for the binary spreadsheet files. No. Then they should have their encodings specified at the command prompt or settings.ini file or get set to the default value. No reason to bother checking this then.
            #    sys.exit( ('Please specify an encoding for: ' + myFileName).encode(consoleEncoding) )

            # Then find the extension of the file.
            myFileNameOnly, myFileExtensionOnly = os.path.splitext(myFileName)

            # if there is no extension, then crash.
            #if myFileExtensionOnly == '':
            #    sys.exit(   ('Warning: Cannot instantiate class using a file that lacks an extension. Reference:\''+myFileName+'\'').encode(consoleEncoding)   ) # Update: Just assume it is a text file instead and import as line-by-line.

            #createOnlyInMemory == True means that the Strawberry() will be created in memory instead of from a file, so do not try to import the contents from a file.
            #createOnlyInMemory == False means import the base contents for Strawberry() from a file.
            # Update: Maybe this should just be implied instead of explicit? Would make for a simpiler UI. Are there any situations where explicit is useful?
            #if createOnlyInMemory == False:
            # Check to make sure file exists.
            # if file does not exist, then crash.
            #if os.path.isfile(myFileName) != True:
            #    sys.exit(('Error: This file does not exist:\''+myFileName+'\'').encode(consoleEncoding))   

            # if the file exists, then read contents from the file.
            if os.path.isfile(myFileName) == True:
                # if extension = .csv, then call importFromCSV(myFileName)
                if myFileExtensionOnly == '.csv':
                    self.importFromCSV( myFileName, myFileNameEncoding=fileEncoding, removeWhitespaceForCSV=removeWhitespaceForCSV, csvDialect=csvDialect )
                elif myFileExtensionOnly == '.xlsx':
                    self.importFromXLSX( myFileName, fileEncoding, sheetNameInWorkbook=spreadsheetNameInWorkbook, readOnlyMode=self.readOnlyMode )
                elif myFileExtensionOnly == '.xls':
                    self.importFromXLS( myFileName, fileEncoding, sheetNameInWorkbook=spreadsheetNameInWorkbook )
                elif myFileExtensionOnly == '.ods':
                    self.importFromODS( myFileName, fileEncoding, sheetNameInWorkbook=spreadsheetNameInWorkbook )
                else:
                    #Else the file must be a text file to instantiate a class with. Only line-by-line parsing is supported.
                    if ( myFileExtensionOnly != '.txt' ) and ( myFileExtensionOnly != '.text' ):
                        print( ('Warning: Attempting to instantiate chocolate.Strawberry() using file with unknown extension:\'' + myFileExtensionOnly + '\' Reading in line-by-line. This is probably incorrect. Reference:\'' + myFileName + '\'').encode(consoleEncoding))
                    self.importFromTextFile( myFileName, fileEncoding,addHeaderToTextFile=self.addHeaderToTextFile)


    def __str__(self):
        #maybe return the headers from the spreadsheet?
        #return str(spreadsheet[1])
        #return 'pie'
        return str(self.getRow(1))


    # Expects a Python list.
    def appendRow( self, newRow ):
        self.spreadsheet.append( newRow )


    #def appendColumn(self, newColumn) #Does not seem to be needed. Data is just not processed that way. Maybe the people who use pandas would appreciate it? Well, if they use pandas, then they should use pandas instead. #notmyproblemyet


    # This sets the value of the cell based upon the cellAddress in the form of 'A4'.
    def setCellValue( self, cellAddress, value ):
        self.spreadsheet[cellAddress]=value


    # This retuns the value of the cell based upon the cellAddress in the form of 'A4'.
    def getCellValue( self, cellAddress ):
        return self.spreadsheet[cellAddress].value


    # Full name of this function is _getCellAddressFromRawCellString, but was shortened for legibility. Edit: Made it longer again.
    # This functions would return 'B5' from: <Cell 'Sheet'.B5>
#    def _getCellAddressFromRawCellString( self, myInputCellRaw ):
        #print('raw cell data='+str(myInputCellRaw))
        #myInputCellRaw=str(myInputCellRaw)
        #Basically, split the string according to . and then split it again according to > to get back only the CellAddress
#        return str(myInputCellRaw).split('.', maxsplit=1)[1].split('>')[0]
        #return [currentRow, currentColumn


    # This function returns a tuple containing 2 strings that represent a row and column extracted from input Cell address
    # such as returning ['5', 'B'] from: <Cell 'Sheet'.B5>   It also works for complicated cases like AB534.
    def _getRowAndColumnFromRawCellString( self, myInputCellRaw ):
        #print('raw cell data='+str(myInputCellRaw))
        #basically, split the string according to . and then split it again according to > to get back only the CellAddress
        myInputCell=str(myInputCellRaw).split('.', maxsplit=1)[1].split('>')[0]
        #myInputCell=self._getCellAddressFromRawCellString(myInputCellRaw)
 
       # https://openpyxl.readthedocs.io/en/stable/api/openpyxl.utils.cell.html
        # So apparently, there is a proper way to do this as openpyxl.utils.cell.coordinate_from_string('AB25') -> ('AB',25).
        column,row=openpyxl.utils.cell.coordinate_from_string( myInputCell )
        # Swap order. Maybe this should be swapped back? Humm.
        return ( str(row), column )

        # Old code:
        index=0
        for i in range(10): #Magic number.
            try:
                int(myInputCell[index:index+1])
                index=i
                break # This break and the assignment above will only execute when the int conversion works.
            except:
                # This will execute if there is an error, like int('A').
                # This will not execute if the int conversion succeds.
                #print('index='+str(index))
                pass
            index+=1
        #currentColumn=myInputCell[:index] # Does not include character in string[Index] because index is after the :
        #currentRow=myInputCell[index:] # Includes character specified by string[index] because index is before the :
        #return [currentRow, currentColumn]
        # Same as above, but faster.
        return [myInputCell[index:], myInputCell[:index]]

    #Example:
    #myRawCell=''
    #for row in mySpreadsheet:
    #    for i in row:
    #        if i.value == 'lots of pies':
    #            print(str(i) + '=' + str(i.value))
    #            myRawCell=i
    #currentRow, currentColumn = spreadsheet._getRowAndColumnFromRawCellString(myRawCell)


    # Returns a list with the contents of the row number specified.
    # Should return None for any blank entry as in: ['pie', None, 'lots of pies']
    def getRow(self, rowNumber):
        #print(rowNumber)
        #return spreadsheet[rowNumber] #returns the raw cell addresses instead of the values.
        #returns the values in a list

        if isinstance( rowNumber, str) == True:
            rowNumber=int(rowNumber)

        myList=[]
        for cell in self.spreadsheet[rowNumber]:
            if debug == True:
                #print( (str(self.spreadsheet[self._getCellAddressFromRawCellString(cell)].value)+',').encode(consoleEncoding),end='')
                print( ( str(cell.value) + ',').encode(consoleEncoding), end='')
            #myList.append(self.spreadsheet[self._getCellAddressFromRawCellString(cell)].value)
            myList.append( cell.value )

        #lengthOfHeader=len( self.spreadsheet[1] )
        #assert( len(myList) == lengthOfHeader )
        assert( len(myList) == len(self.spreadsheet[1]) )

        if debug == True:
            print('')
        return myList


    # Returns a list with the contents of the column specified (by letter).
    # Should return None for any blank entry as in: ['pie', None, 'lots of pies']
    def getColumn(self, columnLetter):

        # https://openpyxl.readthedocs.io/en/stable/api/openpyxl.utils.cell.html
        if isinstance( columnLetter, int) == True:
            # Convert an integer to a column letter (3 -> 'C') so that the calling code does not have to care.
            columnLetter = openpyxl.utils.cell.get_column_letter(columnLetter)

        myList=[]
        # Update: Would the built in iterators also work here? #Yes, but then how does the iterator/code know not to process undesired columns? Would have to process every column until the right one is found. 
        for cell in self.spreadsheet[columnLetter]:
            #print(str(mySpreadsheet[self._getCellAddressFromRawCellString(cell)].value)+',',end='')
            #myList[i]=mySpreadsheet[self._getCellAddressFromRawCellString(cell)].value  #Doesn't work due to out of index error. Use append() method.
            #myList.append( self.spreadsheet[self._getCellAddressFromRawCellString(cell)].value )
            # v4.
            myList.append( cell.value )

        # Attempt 2.
#        for column in self.spreadsheet.iter_cols():
#            for cell in column:
#                if cell.value == searchTerm:
#                    return self._getRowAndColumnFromRawCellString(cell)[0]
#            break

        assert( len(myList) == len(self.spreadsheet['A']) )

        return myList


    # Helper function that changes the data for a row in mySpreadsheet to what is specified in a python List []
    # Note: This is only for modifying existing rows. To add a brand new row, use append:
        #Example: newRow = ['pies', 'lots of pies']
        #mySpreadsheet.append(newRow)
    # The rowLocation specified is the nth rowLocation, not the [0,1,2,3...] row number because rows start with 1.
    def replaceRow( self, rowLocation, newRowList ):
        if debug == True:
            print( str(len(newRowList) ).encode(consoleEncoding))
            print( str(range(len(newRowList)) ).encode(consoleEncoding))
            print( ('newRowList=' + str(newRowList) ).encode(consoleEncoding) )

        for i in range(len(newRowList)):
            #Syntax for assignment is: mySpreadsheet['A4'] = 'pie'
            #mySpreadsheet['A4'] without an assignment returns: <Cell 'Sheet'.A4> 
            #columns begin with 1 instead of 0, so add 1 when referencing the target column, but not the source because source is a python list which are referenced as list[0], list[1], list[2], list[3], etc

            #Was workaround for Syntax error cannot assign value to function call: mySpreadsheet.cell(row=5, column=3)='pies'  
            #spreadsheet[_getCellAddressFromRawCellString(spreadsheet.cell(row=int(rowLocation), column=i+1))]=newRowList[i]

            #A more direct way of doing the same thing is to use .value without () on the cell after the cell reference.
            self.spreadsheet.cell( row=int(rowLocation), column=i+1 ).value=newRowList[i]
        #return myWorkbook

    #Example: replaceRow(7,newRow)


    def replaceColumn( self, columnLetter, newColumnInAList ):
        #So, how to convert a columnLetter into a number or does column='A' also work?
        #Answer column='A' does not work but there are some built in methods.
        #Documentation: https://openpyxl.readthedocs.io/en/stable/api/openpyxl.utils.cell.html
        #x = openpyxl.utils.column_index_from_string('A')   #returns 1 as an int
        #y= openpyxl.utils.get_column_letter(1)   #returns 'A'
        #Example: mySpreadsheet.cell(row=3, column=openpyxl.utils.column_index_from_string('B')).value='pies'
        if isinstance( columnLetter, str) == True:
            try:
                tempColumnNumber=int( columnLetter )
            except:
                tempColumnNumber=openpyxl.utils.column_index_from_string( columnLetter.upper() )
        else:
            # This needs to be an int. Crash if it is not.
            tempColumnNumber=int( columnLetter )

        if debug == True:
            print( ( 'Replacing column \'' + columnLetter + '\' with the following contents:' ).encode(consoleEncoding) )
            print( str( newColumnInAList ).encode(consoleEncoding) )

        for i in range( len(newColumnInAList) ):
            #Syntax for assignment is: mySpreadsheet['A4'] = 'pie''
            #Rows begin with 1, not 0, so add 1 to the reference row, but not to source list since list starts references at 0.
            self.spreadsheet.cell( row=int( i+1 ), column=tempColumnNumber ).value=newColumnInAList[ i ]

    #Example: replaceColumn('B',newColumnList)


    # Return either None if there is no cell with the search term, or the column letter of the cell if it found it. Case and whitespace sensitive search.
    # Aside: To determine the row, the column, or both from the raw cell address, call self._getRowAndColumnFromRawCellString(rawCellAddress)
    def searchHeaders( self, searchTerm ):
        for row in self.spreadsheet.iter_rows():
            for cell in row:
                if cell.value == searchTerm:
                    return self._getRowAndColumnFromRawCellString(cell)[1]
            break
        return None

        # Old code.
        cellFound=None
        for row in self.spreadsheet[1]:
            for i in row:
                if i.value == searchTerm:
                    cellFound=i
                    break
            #if cellFound != None:
            #    print('found')
            #else:
            #    print('notfound')
            break #stop searching after first row  #Hummmmmm.
        if cellFound == None:
            return None
        #Slower.
        #else:
            #myRowNumber, myColumnLetter = self._getRowAndColumnFromRawCellString(cellFound)
        #return myColumnLetter
        return self._getRowAndColumnFromRawCellString(cellFound)[1]   #Faster.

    #Example:
    #cellFound=None
    #isFound=searchHeader(mySpreadsheet,searchTerm)
    #if isFound == None:
    #    print('was not found')
    #else:
    #    print('searchTerm:\"'+searchTerm+'" was found at:'+str(isFound))


    # This searches the first column for the searchTerm and returns None if not found or the row number if it found it. 
    # Case and whitespace sensitive search.
    # This might not be needed anymore because searching the first column is really only necessary when using chocolate.Strawberry() as cache.xlsx and self.searchCache() was implemented to optimize that use case. When processing every entry in the first column, that implies iterating over every entry anyway, so this function to help find a specific entry to process what would not be used. When is it important to find a specific entry, that is possibly a duplicate, to process outside of cache.xlsx?
    def searchFirstColumn(self, searchTerm):
        for column in self.spreadsheet.iter_cols():
            for cell in column:
                if cell.value == searchTerm:
                    return self._getRowAndColumnFromRawCellString(cell)[0]
            break
        return None


    # This returns either [None, None] if there is no cell with the search term, or a list containing the [row, column], the address. Case and whitespace sensitive.
    #To determine the row, the column, or both from the raw cell address, use self._getRowAndColumnFromRawCellString(rawCellAddress)
    def searchSpreadsheet(self, searchTerm):
        for row in self.spreadsheet.iter_rows():
            for cell in row:
                if cell.value == searchTerm:
                    return self._getRowAndColumnFromRawCellString(cell)
        return [None, None]


    # These return either [None,None] if there is no cell with the search term, or a [list] containing the cell row and the cell column (the address in a list). Case insensitive. Whitespace sensitive.
    # To determine the row, the column, or both from the raw cell address, use self._getRowAndColumnFromRawCellString(rawCellAddress)
    def searchRowsCaseInsensitive(self, searchTerm):
        for row in self.spreadsheet.iter_rows():
            for cell in row:
                if isinstance( cell.value, str ):
                    if cell.value.lower() == str(searchTerm).lower():
                        return self._getRowAndColumnFromRawCellString(cell)
        return [None, None]


    def searchColumnsCaseInsensitive(self, searchTerm):
        for column in self.spreadsheet.iter_cols():
            for cell in column:
                if isinstance( cell.value, str ):
                    if cell.value.lower() == str(searchTerm).lower():
                        return self._getRowAndColumnFromRawCellString(cell)
        return [None, None]


    #Give this function a spreadsheet object (subclass of workbook) and it will print the contents of that sheet. #Updated: Moved to Strawberry() class.
    def printAllTheThings(self):
        for row in self.spreadsheet.iter_rows(min_row=1, values_only=True):
            temp=''
            for cell in row:
                temp=temp+','+str(cell)
            print( str(temp[1:]).encode(consoleEncoding) ) # Ignore first comma , in output

    #Old example: printAllTheThings(mySpreadsheet)
    #New syntax: 
    #mySpreadsheet= Strawberry()
    #mySpreadsheet.printAllTheThings()


    # Export spreadsheet to file, write it to the file system, based upon constructor settings, path, and file extension in the path.
    def export(self, outputFileNameWithPath=None, fileEncoding=defaultTextFileEncoding, columnToExportForTextFiles='A'):
        outputFileNameOnly, outputFileExtensionOnly = os.path.splitext( str(outputFileNameWithPath) )
        pathlib.Path( str(pathlib.Path(outputFileNameWithPath).parent) ).mkdir( parents = True, exist_ok = True )
        if outputFileExtensionOnly == '.csv':
            #Should probably try to handle the path in a sane way.
            self.exportToCSV(outputFileNameWithPath, fileEncoding=self.fileEncoding, csvDialect=self.csvDialect)
        elif outputFileExtensionOnly == '.xlsx':
            self.exportToXLSX(outputFileNameWithPath)
        elif outputFileExtensionOnly == '.xls':
            self.exportToXLS(outputFileNameWithPath)
        elif outputFileExtensionOnly == '.ods':
            self.exportToODS(outputFileNameWithPath)
        elif (outputFileExtensionOnly == '.txt') or (outputFileExtensionOnly == '.text'):
            self.exportToTextFile(outputFileNameWithPath, columnToExport=columnToExportForTextFiles, fileEncoding=self.fileEncoding)
        else:
            print( ( 'Warning: Unable to export chocolate.Strawberry() to file with unknown extension of \''+ outputFileExtensionOnly + '\' Full path: '+ str(outputFileNameWithPath) ).encode(consoleEncoding) )


    # Supports line by line parsing only. Header should already be part of text file.
    def importFromTextFile(self, fileNameWithPath, fileEncoding=defaultTextFileEncoding, addHeaderToTextFile=False):
        if addHeaderToTextFile == True:
            self.appendRow( [ 'rawText', 'reserved', 'metadata' ] )
        # Open file as text file with specified encoding and input error handler.
        with open( fileNameWithPath, 'rt', newline='', encoding=fileEncoding, errors=inputErrorHandling ) as myFileHandle:
        # Create a list from every line and append that list to the current spreadsheet.
            fileContents = myFileHandle.read().splitlines() #readlines() does not work right. It returns a single string with lots of \n, so do not bother. read()+splitlines() works as intended.
            for counter,line in enumerate(fileContents):
                if line.strip() != '':
                    self.appendRow( [  line,'', counter ] )


    #columnToExport to export can be a string or an int. if string, then represents name of column. If int, represents the column in the Strawberry() data structure. The int must be converted to a letter before exporting it.
    #if columnToExport == None: then dynamically calculate what should be exported. Only the translated line furthest to the right is valid to export, along with any untranslated lines.
    # Honestly, exporting to text files does not really make sense unless line-by-line mode was enabled. Maybe remove all \n's from the output then? The translated lines should not have them, so just do not reinsert them and remove them from the source untranslated lines if there is no translated line for that row. Is that sane behavior though? Just leave the data alone as-is. Fundamentally, it is the user's responsibility to format their data before they export it, not this function that handles the actual exporting.
    # When is this useful? What is the use case? It always makes more sense to export as .csv right? Otherwise, a specific column will need to be chosen and that should probably be exposed in the CLI. Otherwise, should a mixed mode be supported? Like exporting the right-most entry in the spreadsheet data structure?
    # That is probably the use case that makes the most sense. Translating a plain .txt file and exporting it as a .txt file. Doing spreadsheet -> .txt file exports makes less sense.
    def exportToTextFile(self, fileNameWithPath, columnToExport=None, fileEncoding=defaultTextFileEncoding):
        #print('Hello World'.encode(consoleEncoding))
        totalLengthOfSpreadsheet=len( getColumn('A') )
        if ( columnToExport == None ) and ( totalLengthOfSpreadsheet <=3 ):
            # The user did not translate anything, so just export the extracted data.
            columnToExport='A'
        if isinstance( columnToExport, int ):
            columnToExport=openpyxl.utils.cell.get_column_letter(columnToExport)
        # Is this logic correct? Probably.
        if ( columnToExport != None ) and ( not isinstance( columnToExport, str ) ):
            print( 'Error: Unknown column to export for spreadsheet. Must be a column or None.'+str(type(columnToExport)) )
            return

        with open( fileNameWithPath, 'wt', newline='', encoding=fileEncoding, errors=outputErrorHandling ) as myFileHandle:
            if isinstance( columnToExport, str):
                # then pull the correct column and write out as-is.
                tempColumn=self.getColumn(columnToExport)
                for counter,data in enumerate(tempColumn):
                    if ( self.addHeaderToTextFile == True ) and ( counter+1 == 1 ):
                        # then skip first row.
                        continue
                    # This does not handle new lines in a sane way if there is a new line in data, but whatever. User's problem. If the lines are not formatted correctly, then they should format them correctly instead of or before exporting as a .txt file.
                    myFileHandle.write(data + '\n')
            # elif columnToExport == None:
            else:
                # then the fun begins.
                for rowNumber in range( totalLengthOfSpreadsheet ):
                    if ( self.addHeaderToTextFile == True ) and ( rowNumber+1 == 1 ):
                        # then skip first row.
                        continue
                    tempRow=getRow( rowNumber+1 )
                    tempString=tempRow[0]
                    for counter,cell in enumerate( tempRow ):
                        if ( counter > 2 ) and ( cell != None ) and ( cell != '' ):
                            tempString=cell
                    if tempString == None:
                        print('Unspecified error.')
                        return
                    # This does not handle new lines correctly if there is a new line in tempString.
                    myFileHandle.write(tempString + '\n')
        print( ('Wrote: '+fileNameWithPath).encode(consoleEncoding) )


    #TODO:
    #1) Export an existing spreadsheet to a file.
    #2) Import a file into an existing spreadsheet or dictionary.
    #References/objects are done using workbooks, not the active spreadsheet.
    #Edit: Return value/reference for reading from files should be done by returning a class instance (object) of Strawberry()
    #Strawberry should have its own methods for writing to files of various formats.
    #All files follow the same rule of the first row being reserved for header values and invalid for inputting/outputting actual data.
    def importFromCSV(self, fileNameWithPath, myFileNameEncoding=defaultTextFileEncoding, removeWhitespaceForCSV=True, csvDialect=None):
        print( ('Reading from: '+fileNameWithPath).encode(consoleEncoding) )
        #import languageCodes.csv, but first check to see if it exists
        if os.path.isfile(fileNameWithPath) != True:
            sys.exit(('\n Error. Unable to find .csv file:"' + fileNameWithPath + '"').encode(consoleEncoding))

        #tempWorkbook = openpyxl.Workbook()
        #tempSpreadsheet = tempWorkbook.active
        #tempSpreadsheet = Strawberry()

        # It looks like quoting fields in csv's that use commas , and new
        # lines works but only as double quotes " and not single quotes '
        # Spaces are also preserved as-is if they are within the commas (,) by default, so remove them
        # If spaces are intended to be within the entry, then the user can encapslate them in double quotes
        # Need to test. Even double quotes might not preserve them. Tested: They do not.
        # Could also just say not supported since it is almost certainly an error for hand-written CSV's.
        # Could also have a flag that switches back and forth.
        # Partial solution, added "removeWhitespaceForCSV" function parameter which defaults to True.
        # Reading from dictionaries can be called with the "False" option for maximum flexibility.
        # New problem: How to expose this functionality to user? Partial solution. Just use sensible defaults and have users fix their input.
        #print(inputErrorHandling)
        with open(fileNameWithPath, newline='', encoding=myFileNameEncoding, errors=inputErrorHandling) as myFile: #shouldn't this be codecs.open and with error handling options? codecs seems to be an alias or something? #Edit: Turns out codecs was a relic from python 2 days. Python 3 integrated all of that, so codecs.open is not needed at all anymore.
            # if csvDialect != None.:
                # implement code related to csvDialects here. Default options are unix, excel and excel-tab
            myCsvHandle = csv.reader(myFile)

            for listOfStrings in myCsvHandle:
                if debug == True:
                    print( str(listOfStrings).encode(consoleEncoding) )
                # Clean up whitespace for entities.
                for i in range( len(listOfStrings) ):
                    if removeWhitespaceForCSV == True:
                        listOfStrings[i]=listOfStrings[i].strip()
                    # Fix types.
                    if listOfStrings[i].lower() == 'true':
                        listOfStrings[i]=True
                    elif listOfStrings[i].lower() == 'false':
                        listOfStrings[i]=False
                    elif ( listOfStrings[i].lower() == 'none' ) or ( listOfStrings[i].lower() == '' ):
                        listOfStrings[i]=None
                    # Leave numbers as strings. They should not be processed anyway, so there is no need to mess with them.

                    #tempSpreadsheet.append(listOfStrings)
                    #tempSpreadsheet.appendRow(listOfStrings)


                self.spreadsheet.append(listOfStrings)
        #return tempWorkbook
        if debug == True:
            self.printAllTheThings()


    def exportToCSV(self, fileNameWithPath, fileEncoding=defaultTextFileEncoding,csvDialect=None):
        with open(fileNameWithPath, 'w', newline='', encoding=fileEncoding, errors=outputErrorHandling) as myOutputFileHandle:
            # if csvDialect != None.:
                # implement code related to csvDialects here. Default options are unix, excel and excel-tab
            myCsvHandle = csv.writer(myOutputFileHandle)

            # Get every row for current spreadsheet.
            # For every row, get each item's value in a list.
            # myCsvHandle.writerow(thatList)
            for row in self.spreadsheet.iter_rows(min_row=1, values_only=True):
                tempList=[]
                for cell in row:
                    tempList.append( str(cell) )
                myCsvHandle.writerow(tempList)

        print( ('Wrote: '+fileNameWithPath).encode(consoleEncoding) )


    def importFromXLSX(self, fileNameWithPath, fileEncoding=defaultTextFileEncoding, sheetNameInWorkbook=None, readOnlyMode=False):
        print( ('Reading from: '+fileNameWithPath).encode(consoleEncoding) )
        self.workbook=openpyxl.load_workbook(filename = fileNameWithPath, read_only=readOnlyMode)
        if sheetNameInWorkbook == None:
            self.spreadsheet=self.workbook.active
        else:
            if sheetNameInWorkbook in self.workbook.sheetnames:
                self.spreadsheet=self.workbook[ sheetNameInWorkbook ]
            else:
                self.workbook.create_sheet( title = str(sheetNameInWorkbook) , index=0 )
                self.spreadsheet = self.workbook[ sheetNameInWorkbook ]
        self.spreadsheetName=self.spreadsheet.title


    # https://openpyxl.readthedocs.io/en/stable/optimized.html
    # read_only requires closing the spreadsheet after use.
    def close(self):
        self.workbook.close()


    def exportToXLSX(self, fileNameWithPath, fileEncoding=defaultTextFileEncoding):
        #print('Hello World'.encode(consoleEncoding))
        #Syntax: 
        #theWorkbook.save(filename="myAwesomeSpreadsheet.xlsx")
        self.workbook.save(filename=fileNameWithPath)
        print( ('Wrote: '+fileNameWithPath).encode(consoleEncoding) )


    def importFromXLS(self, fileNameWithPath, fileEncoding=defaultTextFileEncoding, sheetNameInWorkbook=None):
        print('Hello World'.encode(consoleEncoding))
        #print( ('Reading from: '+fileNameWithPath).encode(consoleEncoding) )
        #return workbook


    def exportToXLS(self, fileNameWithPath, fileEncoding=defaultTextFileEncoding):
        print('Hello World'.encode(consoleEncoding))
        #print( ('Wrote: '+fileNameWithPath).encode(consoleEncoding) )


    def importFromODS(self, fileNameWithPath, fileEncoding=defaultTextFileEncoding, sheetNameInWorkbook=None):
        print('Hello World'.encode(consoleEncoding))
        #print( ('Reading from: '+fileNameWithPath).encode(consoleEncoding) )
        #return workbook


    def exportToODS(self, fileNameWithPath, fileEncoding=defaultTextFileEncoding):
        print('Hello World'.encode(consoleEncoding))
        #print( ('Wrote: '+fileNameWithPath).encode(consoleEncoding) )


    # These are methods that try to optimize using chocolate.Strawberry() as cache.xlsx by indexing the first column into a Python dictionary with its associated row number.
    def initializeCache(self):
        # Technically, if using readOnly mode, then a perfect hash table would provide better 'performance', but not clear how to implement that, so do not worry about it.
        #tempDict={}
        # Build index.
        self.index={}
        for counter,entry in enumerate( self.getColumn('A') ):
            #tempDict[entry]=None
            # Skip adding the header.
            if counter == 0:
                continue
            # Otherwise, populate the index based upon the first column. The payload is the source row.
            if ( entry == None ) or ( entry == '' ):
                raise Exception( 'Unable to initalize cache due to None or empty string values in cache.\nTip: Use cache.rebuildCache() to remove the empty items before trying to initializeCache().' )
            self.index[entry]=counter+1
        # last entry = total length of the index since counting starts at 1. Adding 1 would put it out of bounds. # Update: Incorrect. It would be out of bounds if it was pointing to itself, but it is actually pointing to self.spreadsheet which needs the +1 in order for the pointer in the index to point to the correct cell in self.spreadsheet. Otherwise, it ends up pointing to the cell above it resulting in an off by 1 error.
        if len(self.index) != 0:
            self.lastEntry = len(self.index) + 1
        else:
            # There is a special failure case when initializing an empty index with only 0 or 1 entries in the main self.spreadsheet. In that case, self.lastEntry will remain 0 instead of getting incremented by 1. Then, the next time something gets cache.addToCache(), self.lastEntry will be incremented by 1 and return 1 when the correct address is actually 2, assuming a header row is present in the main self.spreadsheet which it always should be. So, increment self.lastEntry from 0 to 1 here.
            self.lastEntry = 1
        # If this fails, then it should check a variable that if set tries to deduplicate the cache.
        try:
            assert( len(self.index)+1 == len(self.getColumn('A')) )
        except:
            print('len(self.index)+1=',len(self.index) )
            print('len(self.getColumn(A))=',len(self.getColumn('A')))
            print( 'Error: Spreadsheet has duplicate items. Cannot use as cache.\nTip: Use cache.rebuildCache() to remove the duplicate items before trying to initializeCache(). Adding new entries while duplicates exist will corrupt the cache.')
            raise

    # Expects a string and searches through the current cache index. Python dictionaries have an O(1) search time, they are hash tables, compared to O(n) search time on Python lists especially when the last list item is being searched for immediately after an append() opperation. Compared to O(n), O(1) is crazy levels of fast, although even O(log n) would have been an improvement.
    def searchCache( self, myString ):
        if myString == None:
            print( 'Warning: Cannot use searchCache to search for myString=None.' )
            return None
        elif not isinstance(myString, str):
            myString=str(myString)

        if myString.strip() == '':
            print( 'Warning: Cannot use searchCache to search for myString=empty string.' )
            return None
        elif myString in self.index.keys():
            return self.index[ myString ]
        else:
            return None


    # accepts a string or a list with a single item? Answer: Just a string.
    # Are there use cases for multiple items? When would a new entry be added together with a value? Would that be when adding both the untranslated entry and translated entry together? How is that implemented? Answer: The only thing known, unless it is computed dynamically, is the currentColumn in the form of a letter, B, C, D, E, F, and the value of the translated/untranslated pairs. There is no way to know which letter corresponds to which column in a list [A, B, C, D, E] without a way to translate that information, and inserting 'None' to all the unused entries would access the unused cells and expand the memory requirements pointlessly. Instead, only accept input as a string, add the string to the openpyxl spreadsheet, add the string to the index, update the self.lastEntry as needed, and return the row number the entry was added.  Also have some code to deal with duplicates added to the cache in a sane way.
    def addToCache( self, myString ):
        if myString == None:
            print( 'Warning: Cannot use addToCache to update myString=None.' )
            return None
        elif not isinstance(myString, str):
            myString=str(myString)

        if myString.strip() == '':
            print( 'Warning: Cannot use addToCache to search for myString=empty string.' )
            return None

        tempSearchResult = self.searchCache(myString)
        if tempSearchResult == None:
            # then add it to the main spreadsheet.
            self.spreadsheet.append( [myString] )

            # Update the self.lastEntry as needed.
            self.lastEntry += 1

            # then add it to the index.
            self.index[myString]=self.lastEntry

            # And return where it was added.
            return self.lastEntry
        else:
            return tempSearchResult


    def rebuildCache(self, coreHeader=None, extraStrawberryToMerge=None):
        # Algorithim for merging (deduplicating) multiple files:
        # Must match: sheet's name (sheet.title, self.spreadsheetNameInWorkbook), and coreHeader
        # coreHeader does not really need to match in terms of being in A1. Could just use A1 from first spreadsheet as coreHeader and then search for it in other spreadsheets. However, it must be present.
        # for every sheet of the same name, sheet.title in all of the workbooks, turn it into database {} where the coreHeader is used as the master key for all values.
        # The values themselves always consist of one dictionary.
        # That one row specific value dictionary contains header=valueForRow mappings for every column except for the coreHeader column.
        # Once every sheet has been turned into a database for that sheet.title, merge all the databases for the same sheet.title
        # With the data finally merged and the backup database{} for a sheet complete, delete any existing sheet.title in the main workbook. Do not delete the existing workbook.
        # Create a new worksheet with the same name.
        # Add the values into the worksheet.
        # delete database
        # move on to next spreadsheet name

        # Obtain header values. Use header in first cell A1 as core index.
        # coreHeader=self.spreadsheet['A1']
        self.index={}
        database={}

        if coreHeader == None:
            coreHeader=self.spreadsheet['A1'].value

        # tempDatabase has keys derived from coreHeader. data itself is the rest of the headers and the data
        #tempDatabase={ key:value, key:value} where the keys are untranslated strings, one string per row, and the values are another dictionary.
        # The key:value pairs in that other dictionary are header=dataForThatRow
        # The actual return value is a tuple (headers, database) which is returned that way to reliably get back the original headers instead of having to sort through the data to derive them.
        tempDatabase = self._getDatabaseFromSpreadsheet( self.spreadsheet, coreHeader )
        if tempDatabase == None:
            print('Unspecified error turning spreadsheet into database while rebuilding cache. No work done. Exiting.')
            return
        assert isinstance(tempDatabase, tuple)

        headersList=tempDatabase[0]
        headers={}
        # This de-duplicates the headers. Not a good idea actually. Duplicate headers are invalid because it is not clear how to process them, so do it anyway and error out earlier than this if there are duplicate headers.
        for i in headersList:
            headers[i]=None
        # The duplicates should have already been removed.
        assert( len(headersList) == len(headers) )
        tempDatabase=tempDatabase[1]

        # This cycles through the data to get the row headers as a dictionary {}, but the row headers will not be added for a particular item if the translation for a particular item, if that exact cell for that untranslatedString, is None. 
        # That means the header for that item will not exist in that particular rowDictionary, but it might exist in other ones. That makes it difficult to derive the total number of headers in the data, the names of those headers, and their default order.
        # As a workaround, adjust the return value to always return the full list of headers and just assume the tempDatabase is not corrupt.
#        for rowDictionary in tempDatabase.values():
#            for header in rowDictionary.keys():
#                headers[header] = None
#            break
#        print('headers=',end='')
#        for header in headers.keys():
#            print( str(header)+',',end='' )
#        print('')

        #if extraStrawberryToMerge != None:
            # The spreadsheetName property must match for the strawberries to be merged.
            #if not self.spreadsheet.spreadsheetName == extraStrawberryToMerge.spreadsheetName:
                #print('Unable to merge due to different spreadsheet names.')
                # return or rebuild index anyway?
            #else:
                # The spreadsheet in the workbook should be returned by the name of the sheet, self.spreadsheet.title
                # But if passing in the strawberry, it is accessed using the spreadsheet variable.
                #strawberryToMergeDatabase = self._getDatabaseFromSpreadsheet( extraStrawberryToMerge.spreadsheet, coreHeader )

                # Doing a generic dict1.update(dict2) will override all conflicting keys with the values in dict2. Basically, it will resolve conflicts, but also delete half the cache for those entries. Not ideal behavior.
                #tempDatabase.update(strawberryToMergeDatabase)

                # To merge two dictionaries where each dictionary has a key={dictionary}, 1) iterate through the keys of the second one. 2) If the key appears in the first one, 3) then pull both value dictionaries, 4) merge them, and 5) if necessary, update the dictionary in the key={dictionary} in the first dictionary.
                # Fancy merge code goes here.

                #Update headersDictionary.
                #for entry in strawberryToMergeDatabase.keys():
                #    for header in entry.keys():
                #        headers[header] = None
                #     break

        # Delete the existing active spreadsheet in the main workbook. Do not delete the existing workbook.
        self.workbook.remove(self.spreadsheet)

        # Create a new worksheet with the same name.
        # This does not seem to be creating the new spreadsheet with the same name as the old spreadsheet. #Update: Fixed.        self.workbook.create_sheet( title = self.spreadsheetName , index=0 )
        #print(self.workbook.sheetnames)
        self.spreadsheet = self.workbook[ self.spreadsheetName ]

        # Add the values into the worksheet.
        # This needs to construct a list [] in the correct order. The first item in the list is the untranslatedEntry and/or the rowDictionary[coreHeader]=value They shoud be the same. The second item is the item specified by the headers dictionary.
        # Add headers.
        #tempList=[]
        #for globalHeader in headers.keys():
        #    tempList.append( globalHeader )
        #self.appendRow(tempList)
        # Add headers faster.
        self.appendRow( headersList )
        # Sanity check.
        assert( coreHeader == self.getCellValue( 'A1' ) == self.spreadsheet[ 'A1' ].value )

        #for every untranslated entry
        for untranslatedEntry,rowDictionary in tempDatabase.items():
            # Sanity check.
            assert( untranslatedEntry == rowDictionary[coreHeader] )
            # Build the correct row in a list with the order of the data specified by the headers.
            tempList = [ untranslatedEntry ]
            for globalHeader in headers.keys():
                # if the data from the coreHeader is added, then the resulting list will have a duplicate, so prevent that.
                if globalHeader != coreHeader:
                    if (globalHeader in rowDictionary):
                        tempList.append( rowDictionary[globalHeader] )
                    else:
                        tempList.append( None )
            self.appendRow(tempList)

        # Delete database.
        del tempDatabase


    # This function takes a spreadsheet and returns a specially formatted Python dictionary. If the key does not appear, it returns None.
    def _getDatabaseFromSpreadsheet(self, mySpreadsheet, key):
        if ( mySpreadsheet == None ) or ( key == None ):
            return None

        headers=[]
        for entry in mySpreadsheet[1]:
            headers.append(entry.value)
        print( ( 'initial headers=' + str(headers) ).encode(consoleEncoding) )
        print( 'len(headers)=', len(headers) )
        if len(headers) == 0:
            return None

        headersDictionary={}
        for i in headers:
            headersDictionary[i]=None
        if len(headersDictionary) != len(headers):
            print( 'Duplicate headers found. Unable to make sense of data. Exiting.' )
            return None

        keyColumnAsNumber=None
        for counter,entry in enumerate(headers):
            if entry == key:
                keyColumnAsNumber=counter+1 #Columns are letters, not numbers, but number 1 will map to column A, 2 to B, and so forth so add 1 to convert the header's index as a list to a column index as a number.
                break
        if keyColumnAsNumber == None:
            print( 'Error: The column header chosen as an index could not be found in the spreadsheet headers: '+str(key) )
            return None

        if keyColumnAsNumber != 1:
            # Then it is column B or similar. Move contents to Column A.
            #Add a new column.
            #print('pie')
            mySpreadsheet.insert_cols(1)
            #Copy values from old column.
            for column in sheet.iter_cols(min_row=keyColumnAsNumber+1, max_row=keyColumnAsNumber+1, values_only=True):
                for counter,cell in enumerate(column):
                    #self.spreadsheet[cellAddress]=value
                    mySpreadsheet[ 'A' + str( counter+1)]=cell
            #Delete old column.
            mySpreadsheet.delete_cols( keyColumnAsNumber+1 )
            # Fix headers.
            headers=[]
            for entry in mySpreadsheet[1]:
                headers.append(entry.value)
            # Sanity checks.
            for counter,entry in enumerate(headers):
                if entry == key:
                    keyColumnAsNumber=counter+1
                    break
            assert( keyColumnAsNumber == 1)
            assert( headers[0] == key == mySpreadsheet['A1'].value )

        tempDatabase={}
        # tempDatabase['headers']=headers #Workaround. Nevermind. Do this properly.
        # Now the coreHeader is always in A1 and columnA values can be used as keys in tempDatabase. Time to create a dictionary of header:value pairs for each row in order to eventually assign
        # tempDatabase[ ColumnARowDataAsKey ]=rowDictionaryContainingHeaderAndValuePairs

        #for i in range( len(headers) ):
        #    tempDictionary={}
        for rowCounter,row in enumerate( mySpreadsheet.iter_rows(values_only=True) ):
            #skip header
            if rowCounter==0:
                continue

            rowKey=mySpreadsheet[ 'A'+ str(rowCounter+1) ].value
            if rowKey == None:
                print( ('Null key found at row '+ str(rowCounter+1) + '.' ).encode(consoleEncoding) )
                continue
            elif rowKey.strip() == '':
                print( ('Empty string key found at row '+ str(rowCounter+1) + '.').encode(consoleEncoding) )
                continue

            if rowKey in tempDatabase.keys():
                print( ('Duplicate key found at row '+ str(rowCounter+1) + ': '+ rowKey).encode(consoleEncoding) )

            tempRowDict={}
            for columnCounter,cell in enumerate( row ):
                if ( cell != None ) and ( cell != '' ):
                    tempRowDict[ headers[columnCounter] ]=cell 

            tempDatabase[ rowKey ]=tempRowDict

        print( 'len(mySpreadsheet[A]) before rebuilding=', len(mySpreadsheet['A']) )
        print( 'len(tempDatabase) after rebuilding=', len(tempDatabase) )

        # This cycles through the data to get the row headers, but the row headers will not be added for a particular item if that item is None. 
        # That means the header for that item will not exist in that particular rowDictionary, but it might exist in other ones. That makes it difficult to derive the total number of headers in the data, the names of those headers, and their default order.
        # As a workaround, adjust the return value to always return the full list of headers and assume the tempDatabase is not corrupt.
#        print( 'tempDatabaseHeaders=',end='')
        #counter=0
#        for rowDictionary in tempDatabase.values():
#            for key in rowDictionary.keys():
#                print(str(key) + ',',end='')
#            print('')
#            break
        # return a tuple of the headers in the dataset, and the raw dataset itself.
        return (headers, tempDatabase)


"""

# TODO: This section.
# Usage examples, assuming this library is in a subfolder named 'resources':
defaultEncoding='utf-8'
myFileName = 'myFile.txt'

#import sys,pathlib
#sys.path.append( str( pathlib.Path( __file__ ).resolve().parent) )
import resources.chocolate as chocolate

spreadsheet=chocolate.Strawberry()

searchCellRow, searchCellColumn = spreadsheet.search


if dealWithEncodingLibraryIsAvailable == True:
    #Update internal library variables to match main program settings.
    dealWithEncoding.debug=debug
    dealWithEncoding.consoleEncoding=consoleEncoding

    inputEncodingType = dealWithEncoding.ofThisFile(myFileName=inputFileName, rawCommandLineOption=command_Line_arguments.inputEncoding, fallbackEncoding=defaultEncodingType)

    #or, to use only positional arguments
    inputEncodingType = dealWithEncoding.ofThisFile(inputFileName, command_Line_arguments.inputEncoding, defaultEncodingType)

    #or, To detect encoding of a file with the chardet library that has already been determined to exist, and does not consider user preferences, fallback encoding, or a return of None from the chardet library:
    inputEncodingType= dealWithEncoding.detectEncoding(inputFileName)
else:
    inputEncodingType=defaultEncoding

print(inputFileName+' will use encoding type: '+inputEncodingType)

"""
