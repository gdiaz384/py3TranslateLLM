#!/usr/bin/env python3
# -*- coding: UTF-8 -*-
"""
py3TranslateLLM.py translates text using Neural net Machine Translation (NMT) and Large Language Models (LLM).

For usage, see py3TranslateLLM.py -h', README.md, and the source code below.

License:
- For the various libraries used, see the Readme for their licenses and project pages.
- This .py file and the libraries under resources/ are C* gdiaz384 and licensed as GNU Affero GPL v3.
- Feel free to use it, modify it, and distribute it to an unlimited extent, but if you distribute binary files of this program outside of your organization, then please make the source code for those binaries available.
- The imperative to make source code available also applies if using this program as part of a server if that server is publically accessible.

"""
#import various libraries that py3TranslateLLM depends on
import argparse                           #Used to add command line options.
import os, os.path                        #Extract extension from filename, and test if file exists.
from pathlib import Path              #Override file in file system with another and create subfolders.
import sys                                     #End program on fail condition.
import io                                        #Manipulate files (open/read/write/close).
#from io import IOBase               #Test if variable is a file object (an "IOBase" object).
#import datetime                          #Used to get current date and time.
from collections import deque    #Used to hold rolling history of translated items to use as context for new translations.
import requests                            #Do basic http stuff, like submitting post/get requests to APIs. Must be installed using: 'pip install requests'
import openpyxl                           #Used as the core internal data structure and also to read/write xlsx files. Must be installed using pip.
import resources.openpyxlHelpers as openpyxlHelpers #A helper/wrapper library to aid in using openpyxl as a datastructure.
import resources.dealWithEncoding as dealWithEncoding   #dealWithEncoding implements the 'chardet' library which is installed with 'pip install chardet'
import resources.py3TranslateLLMfunctions as py3TranslateLLMfunctions  #Moved most generic functions here to increase code readability and enforce function best practices.
#from resources.py3TranslateLLMfunctions import * #Do not use this syntax. Ever. The * is fine, but the 'from' breaks everything because it copies everything instead of pointing to the original resources which makes updating library variables borderline impossible.


#set defaults and static variables
versionString='v0.1 - 2024Jan21 pre-alpha'

defaultTextEncoding='utf-8'
defaultTextEncodingForKSFiles='shift-jis'
defaultConsoleEncodingType='utf-8'

defaultLanguageCodesFile='resources/languageCodes.csv'
defaultSourceLanguage='Japanese'
defaultTargetLanguage='English'
#Valid options are 'English (American)' or 'English (British)'
defaultEnglishLanguage='English (American)'
#Valid options are 'Chinese (simplified)' or 'Chinese (traditional)'
defaultChineseLanguage='Chinese (simplified)'
#Valid options are 'Portuguese (European)' or 'Portuguese (Brazilian)'
defaultPortugueseLanguage='Portuguese (European)'

defaultKoboldCppPort=5001
defaultSugoiPort=14366
#defaultSugoiPort=14467
defaultInputEncodingErrorHandler='strict'
defaultOutputEncodingErrorHandler='namereplace'

defaultLinesThatBeginWithThisAreComments='#'
defaultAssignmentOperatorInSettingsFile='='
defaultMetadataDelimiter='_'
defaultScriptSettingsFileExtension='.ini'

#defaultDebugSetting=False
#defaultDebugSetting=True

translationEngines='parseOnly, koboldcpp, deepl_api_free, deepl_api_pro, deepl_web, fairseq, sugoi'
usageHelp='\n Usage: python py3TranslateLLM --help  Example: py3TranslateLLM -mode KoboldCpp -f myInputFile.ks \n Translation Engines: '+translationEngines+'.'




# So, py3translateLLM.py also supports reading program settings from py3translateLLM.ini in addition to the command prompt.
#Rule is that settings from the command line take #1 precedence
#Then #2) entries from .ini
#Then #3) settings hardcoded in .py
#So, initialize all non-boolean CLI settings to None. Then, only change them using settings.ini or defaultSettings if they are still None after reading from CLI. Any settings that are not None were set using the CLI, so leave them alone.
#For boolean values, they are all False by default. Only change to True if the scriptSettingsDictionary set them to True. Well, that means any CLI settings as intended to be default of 'False' will be overriden to True. Well, that is a user error. Let them deal with it since they were the ones that decided to change the default settings.


#add command line options
commandLineParser=argparse.ArgumentParser(description='Description: CLI wrapper script for various NMT and LLM models.' + usageHelp)
commandLineParser.add_argument('-mode', '--translationEngine', help='Specify translation engine to use, options='+translationEngines+'.',type=str)

commandLineParser.add_argument('-f', '--fileToTranslate', help='Either the raw file to translate or the spreadsheet file to resume translating from, including path.',default=None,type=str)
commandLineParser.add_argument('-fe', '--fileToTranslateEncoding', help='The encoding of the input file. Default='+defaultTextEncoding,default=None,type=str)
commandLineParser.add_argument('-o', '--outputFile', help='The file to insert translations into, including path. Default is same as input file.',default=None,type=str)
commandLineParser.add_argument('-ofe', '--outputFileEncoding', help='The encoding of the output file. Default is same as input file.',default=None,type=str)
commandLineParser.add_argument('-pfile', '--parsingSettingsFile', help='This file defines how to parse raw text and .ks files. It is required for text and .ks files. If not specified, a template will be created.', default=None,type=str)
commandLineParser.add_argument('-pfe', '--parsingSettingsFileEncoding', help='Specify encoding for parsing definitions file, default='+defaultTextEncoding,default=None,type=str)
commandLineParser.add_argument('-p', '--promptFile', help='This file has the prompt for the LLM.', default=None,type=str)
commandLineParser.add_argument('-pe', '--promptFileEncoding', help='Specify encoding for prompt file, default='+defaultTextEncoding,default=None,type=str)

commandLineParser.add_argument('-lcf', '--languageCodesFile', help='Specify a custom name and path for languageCodes.csv. Default=\''+defaultLanguageCodesFile+'\'.',default=None,type=str)
commandLineParser.add_argument('-lcfe', '--languageCodesFileEncoding', help='The encoding of file languageCodes.csv, default='+defaultTextEncoding,default=None,type=str)
commandLineParser.add_argument('-sl', '--sourceLanguage', help='Specify language of source text. Default='+defaultSourceLanguage,default=None,type=str)
commandLineParser.add_argument('-tl', '--targetLanguage', help='Specify language of source text. Default='+defaultTargetLanguage,default=None,type=str)

commandLineParser.add_argument('-cn', '--characterNamesDictionary', help='The file name and path of characterNames.csv',default=None,type=str)
commandLineParser.add_argument('-cne', '--characterNamesDictionaryEncoding', help='The encoding of file characterNames.csv, Default='+defaultTextEncoding,default=None,type=str)
commandLineParser.add_argument('-pred', '--preTranslationDictionary', help='The file name and path of preTranslation.csv',default=None,type=str)
commandLineParser.add_argument('-prede', '--preTranslationDictionaryEncoding', help='The encoding of file preTranslation.csv. Default='+defaultTextEncoding,default=None,type=str)
commandLineParser.add_argument('-postd', '--postTranslationDictionary', help='The file name and path of postTranslation.csv.',default=None,type=str)
commandLineParser.add_argument('-postde', '--postTranslationDictionaryEncoding', help='The encoding of file postTranslation.csv. Default='+defaultTextEncoding,default=None,type=str)
commandLineParser.add_argument('-postwd', '--postWritingToFileDictionary', help='The file name and path of postWritingToFile.csv.',default=None,type=str)
commandLineParser.add_argument('-postwde', '--postWritingToFileDictionaryEncoding', help='The encoding of file postWritingToFile.csv. Default='+defaultTextEncoding,default=None,type=str)

commandLineParser.add_argument('-lbl', '--lineByLineMode', help='Store and translate lines one at a time. Disables grouping lines by delimitor and paragraph style translations.',action='store_true')
commandLineParser.add_argument('-r', '--resume', help='Attempt to resume previously interupted operation. No gurantees.',action='store_true')

commandLineParser.add_argument('-a', '--address', help='Specify the protocol and IP for NMT/LLM server, Example: http://192.168.0.100',default=None,type=str)
commandLineParser.add_argument('--port', help='Specify the port for the NMT/LLM server. Example: 5001',default=None,type=str)

commandLineParser.add_argument('-ieh', '--inputErrorHandling', help='If the wrong input codec is specified, how should the resulting conversion errors be handled? See: docs.python.org/3.7/library/codecs.html#error-handlers Default=\''+defaultInputEncodingErrorHandler+'\'.',default=None,type=str)
commandLineParser.add_argument('-eh', '--outputErrorHandling', help='How should output conversion errors between incompatible encodings be handled? See: docs.python.org/3.7/library/codecs.html#error-handlers Default=\''+defaultOutputEncodingErrorHandler+'\'.',default=None,type=str)

commandLineParser.add_argument('-ce', '--consoleEncoding', help='Specify encoding for standard output. Default='+defaultConsoleEncodingType,default=None,type=str)

commandLineParser.add_argument('-vb', '--verbose', help='Print more information.',action='store_true')
commandLineParser.add_argument('-d', '--debug', help='Print too much information.',action='store_true')
commandLineParser.add_argument('-v', '--version', help='Print version information and exit.',action='store_true')    


#import options from command line options
commandLineArguments=commandLineParser.parse_args()
#print( 'debugSettingFromCLI='+str(commandLineArguments.debug) )

translationEngine=commandLineArguments.translationEngine
fileToTranslate=commandLineArguments.fileToTranslate
parsingSettingsFile=commandLineArguments.parsingSettingsFile
outputFile=commandLineArguments.outputFile
promptFile=commandLineArguments.promptFile

languageCodesFile=commandLineArguments.languageCodesFile
sourceLanguage=commandLineArguments.sourceLanguage
targetLanguage=commandLineArguments.targetLanguage

characterNamesDictionary=commandLineArguments.characterNamesDictionary
preTranslationDictionary=commandLineArguments.preTranslationDictionary
postTranslationDictionary=commandLineArguments.postTranslationDictionary
postWritingToFileDictionary=commandLineArguments.postWritingToFileDictionary

lineByLineMode=commandLineArguments.lineByLineMode
resume=commandLineArguments.resume
address=commandLineArguments.address  #Must be reachable. How to test for that?
port=commandLineArguments.port                #Port should be conditionaly guessed. If no port specified and an address was specified, then try to guess port as either 80, 443, or default settings depending upon protocol and translationEngine selected.

inputErrorHandling=commandLineArguments.inputErrorHandling
outputErrorHandling=commandLineArguments.outputErrorHandling

consoleEncoding=commandLineArguments.consoleEncoding
verbose=commandLineArguments.verbose
debug=commandLineArguments.debug
version=commandLineArguments.version


# Basically, the final value of consoleEncoding is unclear because the settings.ini has not been read yet,
# but it needs to be used immediately for '--version'. Add a temporary console encoding variable.
# If the user specified one at the CLI, use that setting, otherwise use the hardcoded defaults.
if commandLineArguments.consoleEncoding != None:
    tempConsoleEncoding=commandLineArguments.consoleEncoding
else:
    tempConsoleEncoding=defaultConsoleEncodingType
if debug == True:
    print( ('tempConsoleEncoding='+tempConsoleEncoding).encode(tempConsoleEncoding) )
if version == True:
    sys.exit( (versionString).encode(tempConsoleEncoding) )


#Add stub encoding options. All of these are most certainly None, but they need to exist for locals() to find them so they can get updated.
fileToTranslateEncoding=commandLineArguments.fileToTranslateEncoding
outputFileEncoding=commandLineArguments.outputFileEncoding
promptFileEncoding=commandLineArguments.promptFileEncoding

parsingSettingsFileEncoding=commandLineArguments.parsingSettingsFileEncoding
languageCodesFileEncoding=commandLineArguments.languageCodesFileEncoding

characterNamesDictionaryEncoding=commandLineArguments.characterNamesDictionaryEncoding
preTranslationDictionaryEncoding=commandLineArguments.preTranslationDictionaryEncoding
postTranslationDictionaryEncoding=commandLineArguments.postTranslationDictionaryEncoding
postWritingToFileDictionaryEncoding=commandLineArguments.postWritingToFileDictionaryEncoding


# Settings specified at the command prompt take precedence, but py3translateLLM.ini still needs to be pased.
# This function parses that file and places the settings discovered into a dictionary. That dictionary can then be used along with
# the command line options to determine the program settings (prior to validation of those settings).


#This function reads program settings from text files using a predetermined list of rules.
#The text file uses the syntax: setting=value, # are comments, empty/whitespace lines ignored.
#This function builds a dictionary and then returns it to the caller.
#def readSettingsFromTextFile(fileNameWithPath, fileNameEncoding, consoleEncoding=defaultConsoleEncodingType, errorHandlingType=defaultInputEncodingErrorHandler,debug=debug):
#Usage: readSettingsFromTextFile(myfile, myfileEncoding)

currentScriptName = __file__
#currentScriptName = os.path.basename(__file__)  #Returns only the filename of the current script.
#print('file='+__file__) #Returns filename and also the path sometimes.
#This might error out if there is no extension, or maybe it will just return an empty string for the extension.
currentScriptNameOnly, currentScriptExtensionOnly = os.path.splitext(currentScriptName)
scriptSettingsFile=currentScriptNameOnly+defaultScriptSettingsFileExtension

scriptSettingsDictionary=None
#if exist scriptSettingsFile
if os.path.isfile(scriptSettingsFile) == True:
    print( ('Settings file found. Reading settings from: '+scriptSettingsFile).encode(tempConsoleEncoding) )
    scriptSettingsDictionary=py3TranslateLLMfunctions.readSettingsFromTextFile(scriptSettingsFile,defaultTextEncoding, consoleEncoding=tempConsoleEncoding) #Other settings can be specified, but are basically completely unknown at this point, so just use hardcoded defaults instead.


if scriptSettingsDictionary != None:
    # For every entry in dictionary, if the variable referenced by the key's value (current variable's value in script) == None, then set key equal to dictionary's.
    # There is no way to say the above in Python. Batch yes, Python no.
    # Looping through the dictionary /might/ work if this dictionary was read first and then the CLI was allowed to blindly override the values, but that would require changing the order and coming up with yet another fix for --debug flag not working.
#    for x, y in scriptSettingsDictionary.items():
#        print(x)
#        if x == None:
#            print(pie)
    #Since looping does not work, then just dumbly go through each value one by one. This also means the user cannot add arbitrary values or change other settings than the settings explicitly specified here. However, this creates an arbitrary dependency on the .ini file. If the .ini file exists, then every CLI option must also be present there, or attempting to access values not present gives a KeyError.
#    if translationEngine == None:
#        translationEngine = scriptSettingsDictionary['translationEngine']
#    if commandLineArguments.fileToTranslateEncoding == None:
#        #basically, If an encoding was specified at the command prompt, do nothing. But if an encoding was not specified, then just blindly set it to whatever it was set to in the dictionary. No, also wrong. Need to check if dictionary has the entry first.
#        commandLineArguments.fileToTranslateEncoding = scriptSettingsDictionary['fileToTranslateEncoding']
#   Okay, try looping method again.

    #Proxy the dictionary to avoid calling locals() multiple times per loop and in a loop.
    tempDict = locals()
    for x, y in scriptSettingsDictionary.items():
        #if y != None:
        if debug == True:
            print( ('Processing Key: '+str(x)+'='+str(y)+' Current value='+str(tempDict[x]) ).encode(tempConsoleEncoding) )

        #So... locals() returns a dictionary containing all of the currently available variables in the script and their values.
        #That dictionary can then be accessed like any other dictionary, as in myDict['key'], but key has to be a string value already present in the dictionary. Is that because it returns the value of the key? myDictionary['key']='value' works fine for assignment, but not here.
        #In other words, the values accessed this way must already exist, regardless of doing a comparison or assignment operator for them.

        #print (locals()[x])  #This prints the value of the script's variable name 'x', with the name of x derived from the .ini file.
        #Which means that value, the current value in the script, can be compared with None, and only update the variable name conditionally.
        #Since the variable name referenced by x already exists in the module/global scope, but not locally, this is not considered a local assignment.
        #And since it is just getting a value from a dictionary, it is valid to assign values after it. Invoking a function without a sub property on the left side of an assignment always results in an error: function()=value   #Errors out, but this still works: local()[x] = y
        #if locals()[x] == None:
        #    local()[x] = y
        if tempDict[x] == None:
            #Do not bother updating if the new value would be None too.
            if y != None:
                if (verbose == True) or (debug == True):
                    print( ('Updating variable: '+str(x)+' to \''+str(y)+'\'').encode(tempConsoleEncoding) )
                tempDict[x] = y


# if consoleEncoding is still None after reading both the CLI and the settings.ini, then just set it to the default value.
if consoleEncoding == None:
    consoleEncoding=defaultConsoleEncodingType
# if version was specified in the .ini, then print out version string and exit.
if version == True:
    sys.exit( (versionString).encode(consoleEncoding) )
if debug == True:
    print( ('translationEngine='+str(translationEngine)).encode(consoleEncoding) )
    print( ('fileToTranslateEncoding='+str(fileToTranslateEncoding)).encode(consoleEncoding) )

#The variable names needed to be exact for the local()[x] dictionary shenanigans to work, but now that they have been updated, rename them to be more descriptive. The variable names for encoding options will be fixed later.
fileToTranslateFileName=fileToTranslate
parseSettingsFileName=parsingSettingsFile
outputFileName=outputFile
promptFileName=promptFile

languageCodesFileName=languageCodesFile
sourceLanguageRaw=sourceLanguage
targetLanguageRaw=targetLanguage

charaNamesDictionaryFileName=characterNamesDictionary
preDictionaryFileName=preTranslationDictionary
postDictionaryFileName=postTranslationDictionary
postWritingToFileDictionaryFileName=postWritingToFileDictionary


# Now that command line options and .ini have been parsed, update settings for imported libraries.
dealWithEncoding.verbose=verbose
dealWithEncoding.debug=debug
dealWithEncoding.consoleEncoding=consoleEncoding

openpyxlHelpers.verbose=verbose
openpyxlHelpers.debug=debug
openpyxlHelpers.consoleEncoding=consoleEncoding
openpyxlHelpers.inputErrorHandling=inputErrorHandling
openpyxlHelpers.outputErrorHandling=outputErrorHandling

py3TranslateLLMfunctions.verbose=verbose
py3TranslateLLMfunctions.debug=debug
py3TranslateLLMfunctions.consoleEncoding=consoleEncoding
py3TranslateLLMfunctions.inputErrorHandling=inputErrorHandling
py3TranslateLLMfunctions.outputErrorHandling=outputErrorHandling
py3TranslateLLMfunctions.linesThatBeginWithThisAreComments=defaultLinesThatBeginWithThisAreComments
py3TranslateLLMfunctions.assignmentOperatorInSettingsFile=defaultAssignmentOperatorInSettingsFile
py3TranslateLLMfunctions.metadataDelimiter=defaultMetadataDelimiter



# Start to validate input settings and input combinations from parsed imported command line option values.
# Some values need to be set to hardcoded defaults if they were not specified at the command prompt or in settings.ini.
# Others must be present, like certain files.

# Verify translationEngine
mode=None
implemented=False
#validate input from command line options
if translationEngine == None:
    sys.exit( ('Error: Please specify a translation engine. '+usageHelp).encode(consoleEncoding) )
elif (translationEngine.lower()=='parseonly'):
    mode='parseOnly'
    implemented=True
elif (translationEngine.lower()=='koboldcpp'):
    mode='koboldcpp'
    implemented=True
elif (translationEngine.lower()=='deepl_api_free') or (translationEngine.lower()=='deepl-api-free'):
    mode='deepl_api_free'
    #import deepl
elif (translationEngine.lower()=='deepl_api_pro') or (translationEngine.lower()=='deepl-api-pro'):
    mode='deepl_api_pro'
elif (translationEngine.lower()=='deepl_web') or (translationEngine.lower()=='deepl-web'):
    mode='deepl_web'
elif (translationEngine.lower()=='fairseq'):
    mode='fairseq'
elif (translationEngine.lower()=='sugoi') :
    mode='sugoi'
    #Unlike fairseq, Sugoi has a default port association and only supports Jpn->Eng translations, so having a dedicated entry for it is still useful for input validation. However, this mode should be updated to fairseq once input has been validated properly.
else:
    sys.exit(('\n Error. Invalid translation engine specified: "' + translationEngine + '"' + usageHelp).encode(consoleEncoding))

print(('Mode is set to: \''+str(mode)+'\'').encode(consoleEncoding))
if implemented == False:
    sys.exit(('\n\"'+mode+'\" not yet implemented. Please pick another translation engine. \n Translation engines: '+ translationEngines).encode(consoleEncoding))


# Certain files must always exist, like fileToTranslateFileName and languageCodesFileName.
# parseSettingsFileName is only needed if read or writing to text files. Reading from text files is easy to check.
# But how to check if writting to them? If output is .txt, .ks, .ts, then writting to text file. Output can also be based upon input. For output, parseOnly must not be specified, so this output text file check should not be checked with the parseOnly block. Alternatively: The only time parseSettingsFileName is not needed is when writting to output files.
#Errors out if myFile does not exist.
"""
def verifyThisFileExists(myFile,nameOfFileToOutputInCaseOfError=None):
#Usage:
py3TranslateLLMfunctions.verifyThisFileExists('myfile.csv','myfile.csv')
py3TranslateLLMfunctions.verifyThisFileExists(myVar, 'myVar')
"""

py3TranslateLLMfunctions.verifyThisFileExists(fileToTranslateFileName,'fileToTranslateFileName')
py3TranslateLLMfunctions.verifyThisFileExists(languageCodesFileName,'languageCodesFileName')


# Either a raw.unparsed.txt must be specified or a raw.untranslated.csv if selecting one of the other engines.

# Check for NMT without address option. If NMT with address, then handle port assignment. Warn if defaulting to specific port.

#Spreadsheet formats are not valid with parseOnly command, but some valid file must be specified.
#Therefore, that file must always have an extension due to overloading of --inputFile. Not necessarily. parseOnly means the extension could be non-existant but input is still clarified as 100% a text file. Therefore, #split extension code should retun None to fileToTranslateFileNameExtensionOnly in those cases, and that not signify an error during parseOnly. Question: how does os.path.splitext(fileName) actually work? Answer: If the extension does not exist, then it returns an empty string '' object for the extension. A None comparison will not work, but...   if myFileExtOnly == '':   ... will return true and conditionally execute.

#this top comparison, fileToTranslateFileName == None, should never trigger because fileToTranslateFileName is a required input for this program and the argparse library will exit() if it is not present #Edit: Changed to read program settings from .ini, so fileToTranslateFileName will now not always be specified at the command line. It might be None now. #Edit (again): moved this check to function verifyThisFileExists(myFile,nameOfFileToOutputInCaseOfError)
#if fileToTranslateFileName == None:
#    sys.exit(('\n Error: Please specify a valid input file. \n' + usageHelp).encode(consoleEncoding))


fileToTranslateFileNameOnly, fileToTranslateFileNameExtensionOnly = os.path.splitext(fileToTranslateFileName)


#if using parseOnly, a valid file (raw.unparsed.txt and parseDefinitionsFile.txt) must exist. 
if mode == 'parseOnly':
    #check if valid parsing definition file exists 'parseKirikiri.txt'
    py3TranslateLLMfunctions.verifyThisFileExists(parseSettingsFileName,'parseSettingsFileName')
#    if parseSettingsFileName != None:
#        if os.path.isfile(parseSettingsFileName) != True:
#            sys.exit(('\n Error: Unable to find input file "' + fileToTranslateFileName + '"\n' + usageHelp).encode(consoleEncoding))
#    elif parseSettingsFileName == None:
#        print('\n Error: A pasing definitions file is required for \'parseOnly\' mode. Please specify a valid pasing file with --parsingSettingsFile (-pfile) .')#Should not error out even without encode(consoleEncoding) since it is using pure ascii.
#        sys.exit((usageHelp).encode(consoleEncoding))
#    else:
#        sys.exit(' Unspecified error.')
    #check to make sure parseOnly was not called with a spreadsheet extension, as those are invalid combinations
    if (fileToTranslateFileNameExtensionOnly == '.csv') or (fileToTranslateFileNameExtensionOnly == '.xlsx') or (fileToTranslateFileNameExtensionOnly == '.xls') or (fileToTranslateFileNameExtensionOnly == '.ods'):
        sys.exit( ('parseOnly is only valid with text files. It is not valid with spreadsheets: '+str(fileToTranslateFileName)).encode(consoleEncoding) )


# If using koboldcpp, an address must be specified,
# and a prompt file must be specified when using LLMs.
if mode == 'koboldcpp':
    if address == None:
        sys.exit( ('An address must be specified if using koboldcpp.').encode(consoleEncoding) )
# if port not specified, set port to kobold default port
    if port == None:
        print( ('Warning: No port specified for koboldcpp. Using default port of'+str(defaultKoboldCppPort)).encode(consoleEncoding) )
        port=defaultKoboldCppPort
    elif port != None:
        try:
            port=int(port)
            assert port > 0
            assert port <= 65535
        except:
            sys.exit( ('Unable to verify port number: \''+str(port)+'\' Must be 1-65535.').encode(consoleEncoding) )
    py3TranslateLLMfunctions.verifyThisFileExists(promptFileName, 'promptFileName')

# if using deepl_api_free, pro
if (mode == 'deepl_api_free') or (mode == 'deepl_api_pro'):
# library must be available
    try:
        import deepl
    except:
        sys.exit( ('DeepL\'s python library is not available. Please install using: pip install deepl').encode(consoleEncoding) )
# api key must exist
    #maybe check both environmental variable and also a file?
    #The deepL library also requires it, so how does it need to be specified? Does it check any environmental variable for it?


# if using deepl_web... TODO validate anything it needs
    #probably the external chrome.exe right? Maybe requests library and scraping library.
    #Might need G-Chrome, + chromedriver.exe, and also a library wrapper, like Selenium.
    #Could require Chromium to be downloaded to resources/chromium/[platform]/chrome[.exe]


# if using fairseq, address must be specified
if (mode == 'fairseq'):
    if address == None:
        sys.exit( ('Error: Please specify address for fairseq translation engine. Example: http://localhost').encode(consoleEncoding) )
    if port == None:
        # try to guess port from protocol and warn user
        #split address using : and return everything before:
        protocol=address.split(':')[0]
        if protocol.lower() == 'http':
            port=80
        elif protocol.lower() == 'https':
            port=443
        else:
            sys.exit( ('Port not specified and unable to guess port from protocol \''+protocol+'\' of address \''+address+'\' Please specify a valid port number between 1-65535.').encode(consoleEncoding) )
        print( ('Warning: No port was specified. Defaulting to port \''+port+'\' based on protocol \''+protocol+'\'. This is probably incorrect.') )
    elif port != None:
        try:
            port=int(port)
            assert port > 0
            assert port <= 65535
        except:
            sys.exit( ('Unable to verify port number: \''+str(port)+'\' Must be 1-65535.').encode(consoleEncoding) )


# if using sugoi, address must be specified.
# if port not specified, set port to default sugoi port and warn user.
# Change mode to fairseq at the end.
if (mode == 'sugoi'):
    if address == None:
        sys.exit( ('Error: Please specify address for sugoi translation engine. Example: http://localhost').encode(consoleEncoding) )
    if port == None:
        print( ('Warning: No port specified for sugoi translation engine. Using default port of: '+str(defaultSugoiPort)).encode(consoleEncoding) )
        port=defaultSugoiPort
    elif port != None:
        try:
            port=int(port)
            assert port > 0
            assert port <= 65535
        except:
            sys.exit( ('Unable to verify port number: \''+str(port)+'\' Must be 1-65535.').encode(consoleEncoding) )
    mode='fairseq'




if sourceLanguageRaw == None:
    sourceLanguageRaw = defaultSourceLanguage
if targetLanguageRaw == None:
    targetLanguageRaw = defaultTargetLanguage

#Substitute a few language entries to certain defaults due to collisions and aliases.
if sourceLanguageRaw.lower() == 'english':
    sourceLanguageRaw = 'English (American)'
elif sourceLanguageRaw.lower() == 'castilian':
    sourceLanguageRaw = 'Spanish'
elif sourceLanguageRaw.lower() == 'chinese':
    sourceLanguageRaw = 'Chinese (simplified)'
elif sourceLanguageRaw.lower() == 'portuguese':
    sourceLanguageRaw = 'Portuguese (European)'

if targetLanguageRaw.lower() == 'english':
    targetLanguageRaw = 'English (American)'
elif targetLanguageRaw.lower() == 'castilian':
    targetLanguageRaw='Spanish'


#Set encodings Last
##Set Encodings for files using dealWithEncoding.py helper library as dealWithEncoding.ofThisFile(myfile). 

#First one is a special case since the encoding must be changed to shift-jis in the case of .ks (file extension).
#If an encoding was not specified for inputFile, and the extension is .ks, then default to shift-jis encoding and warn user of change.
#if an input file name was specified, then
if fileToTranslateFileName != None:
    #fileToTranslateFileNameOnly, fileToTranslateFileNameExtensionOnly = os.path.splitext(fileToTranslateFileName)#Moved further up, made global/module wide. #Edit, moved into a function that is always called. Also checks for == None condition and errors out the program if not specified.
    #print(fileToTranslateFileNameOnly)
    #print('Extension:'+fileToTranslateFileNameExtensionOnly)
    #if no encoding was specified, then...
    if commandLineArguments.fileToTranslateEncoding == None:
        if fileToTranslateFileNameExtensionOnly == '.ks':
            fileToTranslateEncoding=defaultTextEncodingForKSFiles #set encoding to shift-jis
            print(('Warning: KAG3 (.ks) input file specified, but no encoding was specified. Defaulting to '+defaultTextEncodingForKSFiles+' instead of '+defaultTextEncoding+'. If this behavior is not desired or produces corrupt input, please specify an encoding using --fileToTranslateEncoding (-fe) option instead.').encode(consoleEncoding))
        else:
            #If a file was specified, and if an encoding was not specified, and if the file is not a .ks file, then try to detect encoding
            #fileToTranslateEncoding=defaultTextEncoding#set encoding to default encoding
            fileToTranslateEncoding = dealWithEncoding.ofThisFile(fileToTranslateFileName, fileToTranslateEncoding, defaultTextEncoding)
    else:
        fileToTranslateEncoding=commandLineArguments.fileToTranslateEncoding#set encoding to user specified encoding
else:
    fileToTranslateEncoding=defaultTextEncoding#if an input file name was not specified, set encoding to default encoding

#outputFileEncoding=
#For the output file encoding:  
#if the user specified an input file, use the input file's encoding for the output file
#if the user did not specify an input file, then the program cannot run, but they might have specified a spreadsheet (.csv, .xlsx, .xls, .ods)
#.csv's work normally since those are both plaintext and spreadsheets but reading encoding from .xlsx, .xls, and .ods is not possible without either hardcoding a specific encoding or using the designated libraries. dealWithEncoding should deal with it. If asked what the encoding of a binary spreadsheet format, then just return what the user specified at the command prompt or the default/fallbackEncoding.

#set rest of encodings using dealWithEncoding.ofThisFile(myFileName, rawCommandLineOption, fallbackEncoding):

#parsingSettingsFileEncoding=commandLineArguments.parsingSettingsFileEncoding
parseSettingsFileEncoding = dealWithEncoding.ofThisFile(parseSettingsFileName, parsingSettingsFileEncoding, defaultTextEncoding)

promptFileEncoding = dealWithEncoding.ofThisFile(promptFileName, promptFileEncoding, defaultTextEncoding)

#languageCodesEncoding=commandLineArguments.languageCodesFileEncoding
languageCodesEncoding = dealWithEncoding.ofThisFile(languageCodesFileName, languageCodesFileEncoding, defaultTextEncoding)

#charaNamesDictionaryEncoding=commandLineArguments.characterNamesDictionaryEncoding
charaNamesDictionaryEncoding = dealWithEncoding.ofThisFile(charaNamesDictionaryFileName, characterNamesDictionaryEncoding, defaultTextEncoding)

#preDictionaryEncoding=commandLineArguments.preTranslationDictionaryEncoding
preDictionaryEncoding = dealWithEncoding.ofThisFile(preDictionaryFileName, preTranslationDictionaryEncoding, defaultTextEncoding)

#postDictionaryEncoding=commandLineArguments.postTranslationDictionaryEncoding
postDictionaryEncoding = dealWithEncoding.ofThisFile(postDictionaryFileName, postTranslationDictionaryEncoding, defaultTextEncoding)

#postWritingToFileDictionaryEncoding=commandLineArguments.postWritingToFileDictionaryEncoding
postWritingToFileDictionaryEncoding = dealWithEncoding.ofThisFile(postWritingToFileDictionaryFileName, postWritingToFileDictionaryEncoding, defaultTextEncoding)


#Now that input validation is /finally/ done, define functions.
###### Define various helper functions ###### 





#TODO:
#This function processes raw data (.ks, .txt. .ts) using a parse file. The extracted data is meant to be loaded into the main workbook data structure for further processing.
def importFromRawTextFile(fileNameWithPath, fileNameEncoding, parseFile, parseFileEncoding):
    #print('Hello World'.encode(consoleEncoding))
    global mainSpreadsheet
    #fileNameWithPath does not need an extension check because it is already clear
    #Check to make sure fileNameWithPath exists.
    #Check to make sure parseFile exists.



##### End Functions list #####


#read in parseFile which returns as a dictionary. Parse file is usually needed because it has both parse settings (start of processing) and wordWrap settings (end of processing). However, neither parsing nor wordWrap are always needed since user can specify a seperate outfile which just dumps mainSpreadsheet with the translated values.
parseSettings=None
if parseSettingsFileName != None:
    parseSettings=py3TranslateLLMfunctions.readSettingsFromTextFile(parseSettingsFileName,parseSettingsFileEncoding,consoleEncoding=consoleEncoding,errorHandlingType=inputErrorHandling,debug=debug)


if (verbose == True) or (debug == True):
    print( ('verbose='+str(verbose)).encode(consoleEncoding) )
    print( ('debug='+str(debug)).encode(consoleEncoding) )
    print( sourceLanguageRaw.encode(consoleEncoding) )
    print( targetLanguageRaw.encode(consoleEncoding) )

#Instantiate basket of Strawberries. Start with languageCodes.csv Edit: Use Python dictionaries instead for the dictionary.csv files.
#languageCodes.csv could also be a dictionary or multidimension array, but then searching through it would be annoying, so leave as a Strawberry.
#Read in and process languageCodes.csv
#Format specificiation for languageCodes.csv
#Name of language in English, ISO 639 Code, ISO 639-2 Code
#https://www.loc.gov/standards/iso639-2/php/code_list.php
#Language list (Mostly) only languages supported by DeepL are currently listed, case insensitive.
#Syntax:
#The first row is entirely headers (column labels). Entries start from the 2nd row (row[1] if python list, row[2] if spreadsheet data structure):
#row 0) [languageNameReadable, 
#row 1) TwoLetterLanguageCodeThatIsNotAlwaysTwoLetters,
#row 2) ThreeLetterLanguageCodeThatIsNotAlwaysThreeLetters,
#row 3) DoesDeepLSupportThisLanguageTrueOrFalse,
#row 4) DoesThisLanguageNeedACustomSourceLanguageTrueOrFalse (for single source language but many target language languages like EN->EN-US/EN-GB)
#row 5) If #4 is True, then the full name source language
#row 6) If #4 is True, then the two letter code of the source language
#row 7) If #4 is True, then the three letter code of the source language
#Examples for 5-7: English, EN, ENG; Portuguese, PT-PT, POR
#Each row/entry is/has eight columns total.

#languageCodesWorkbook = openpyxlHelpers.importFromCSV(languageCodesFileName)
#languageCodesSpreadsheet = languageCodesWorkbook.active

languageCodesSpreadsheet=openpyxlHelpers.Strawberry(languageCodesFileName,languageCodesEncoding,ignoreWhitespaceForCSV=True)

#replace this code with dedicated search functions from openpyxlHelpers
#use....openpyxlHelpers.searchColumnsCaseInsensitive(spreadsheet, searchTerm)
#languageCodesSpreadsheet

#sourceLanguageCellRow, sourceLanguageCellColumn = languageCodesSpreadsheet.searchColumnsCaseInsensitive('lav')
#sourceLanguageCellRow, sourceLanguageCellColumn = languageCodesSpreadsheet.searchColumnsCaseInsensitive('japanese')
sourceLanguageCellRow, sourceLanguageCellColumn = languageCodesSpreadsheet.searchColumnsCaseInsensitive(sourceLanguageRaw)

if (sourceLanguageCellRow == None) or (sourceLanguageCellColumn == None):
    sys.exit( ('Error: Unable to find source language \''+str(sourceLanguageRaw)+'\' in file: '+ str(languageCodesFileName)).encode(consoleEncoding) )

print( ('Using sourceLanguage \''+sourceLanguageRaw+'\' found at \''+str(sourceLanguageCellColumn)+str(sourceLanguageCellRow)+'\' of: '+languageCodesFileName).encode(consoleEncoding) )


sourceLanguageFullRow = languageCodesSpreadsheet.getRow(sourceLanguageCellRow)
if debug == True:
    print( str(sourceLanguageFullRow).encode(consoleEncoding) )

internalSourceLanguageName=sourceLanguageFullRow[0]
internalSourceLanguageTwoCode=sourceLanguageFullRow[1]
internalSourceLanguageThreeCode=sourceLanguageFullRow[2]
if debug == True:
    print( ('internalSourceLanguageName='+internalSourceLanguageName).encode(consoleEncoding) )
    print( ('internalSourceLanguageTwoCode='+internalSourceLanguageTwoCode).encode(consoleEncoding) )
    print( ('internalSourceLanguageThreeCode='+internalSourceLanguageThreeCode).encode(consoleEncoding) )


#targetLanguageCellRow, targetLanguageCellColumn = languageCodesSpreadsheet.searchColumnsCaseInsensitive('pie')
targetLanguageCellRow, targetLanguageCellColumn = languageCodesSpreadsheet.searchColumnsCaseInsensitive(targetLanguageRaw)
if (targetLanguageCellRow == None) or (targetLanguageCellColumn == None):
    sys.exit( ('Error: Unable to find target language \''+targetLanguageRaw+'\' in file: '+ languageCodesFileName).encode(consoleEncoding) )

print( ('Using targetLanguage \''+targetLanguageRaw+'\' found at \''+str(targetLanguageCellColumn)+str(targetLanguageCellRow)+'\' of: '+languageCodesFileName).encode(consoleEncoding) )

targetLanguageFullRow = languageCodesSpreadsheet.getRow(targetLanguageCellRow)
if debug == True:
    print( str(targetLanguageFullRow).encode(consoleEncoding) )

internalDestinationLanguageName=targetLanguageFullRow[0]
internalDestinationLanguageTwoCode=targetLanguageFullRow[1]
internalDestinationLanguageThreeCode=targetLanguageFullRow[2]
if debug == True:
    print( ('internalDestinationLanguageName='+internalDestinationLanguageName).encode(consoleEncoding) )
    print( ('internalDestinationLanguageTwoCode='+internalDestinationLanguageTwoCode).encode(consoleEncoding) )
    print( ('internalDestinationLanguageThreeCode='+internalDestinationLanguageThreeCode).encode(consoleEncoding) )






sys.exit('end reached')


"""
#if debug == True:
#    print('')
#    print(('The following has been read from ' + str(languageCodesFileName)).encode(consoleEncoding))
#    printAllTheThings(languageCodesSpreadsheet)

#Search for valid sourceLanguageRaw and targetLanguageRaw. Case insensitive.
for column in languageCodesSpreadsheet.iter_cols():
    for cell in column:
        if isinstance(cell.value, str):
            if cell.value.lower() == sourceLanguageRaw.lower():
                sourceLanguageCell=cell
                break
            #else:
                #cellVal=cell.value.lower()
                #print(cellVal)
                #srcLang=sourceLanguageRaw.lower()
                #print(srcLang)
                #print('"'+cellVal+'"!="'+srcLang+'"')

if sourceLanguageCell == None:
    sys.exit(('\n Error. The following language was not found in"' + languageCodesFileName + '":"' + sourceLanguageRaw+'"').encode(consoleEncoding))
else:
    print(('SourceLanguage:'+sourceLanguageRaw+':'+str(sourceLanguageCell)).encode(consoleEncoding))
#Assume sourceLanguageCell has a valid value now.
sourceLanguageRow, sourceLanguageColumn = getRowAndColumnFromCell(sourceLanguageCell)
#print(languageCodesSpreadsheet['A'+sourceLanguageRow].value)

if debug == True:
    print((str(languageCodesSpreadsheet['A'+sourceLanguageRow].value)).encode(consoleEncoding))
    print((str(languageCodesSpreadsheet['B'+sourceLanguageRow].value)).encode(consoleEncoding))
    print((str(languageCodesSpreadsheet['C'+sourceLanguageRow].value)).encode(consoleEncoding))
    print((str(languageCodesSpreadsheet['D'+sourceLanguageRow].value)).encode(consoleEncoding))
    print((str(languageCodesSpreadsheet['E'+sourceLanguageRow].value)).encode(consoleEncoding))
    print((str(languageCodesSpreadsheet['F'+sourceLanguageRow].value)).encode(consoleEncoding))
    print((str(languageCodesSpreadsheet['G'+sourceLanguageRow].value)).encode(consoleEncoding))
    print((str(languageCodesSpreadsheet['H'+sourceLanguageRow].value)).encode(consoleEncoding))

if str(languageCodesSpreadsheet['E'+sourceLanguageRow].value) == 'False':
    internalSourceLanguageName=languageCodesSpreadsheet['A'+sourceLanguageRow].value
    internalSourceLanguageTwoCode=languageCodesSpreadsheet['B'+sourceLanguageRow].value
    internalSourceLanguageThreeCode=languageCodesSpreadsheet['C'+sourceLanguageRow].value
elif str(languageCodesSpreadsheet['E'+sourceLanguageRow].value) == 'True':
    internalSourceLanguageName=languageCodesSpreadsheet['F'+sourceLanguageRow].value
    internalSourceLanguageTwoCode=languageCodesSpreadsheet['G'+sourceLanguageRow].value
    internalSourceLanguageThreeCode=languageCodesSpreadsheet['H'+sourceLanguageRow].value
else:
    print('')
    print((languageCodesSpreadsheet['E'+sourceLanguageRow].value).encode())
    sys.exit('Unspecified error.'.encode(consoleEncoding))

if debug == True:
    print(str(bool(languageCodesSpreadsheet['E'+sourceLanguageRow].value)).encode())
    print(internalSourceLanguageName.encode())
    print(internalSourceLanguageTwoCode.encode())
    print(internalSourceLanguageThreeCode.encode())

targetLanguageCell=None
#print(sourceLanguageRaw)

#targetLanguageCell = searchSpreadsheet(languageCodesSpreadsheet,targetLanguageRaw) #case sensitive, do not use

#Search for valid sourceLanguageRaw and targetLanguageRaw. Case insensitive.
for column in languageCodesSpreadsheet.iter_cols():
    for cell in column:
        if isinstance(cell.value, str):
            if cell.value.lower() == targetLanguageRaw.lower():
                targetLanguageCell=cell
                break
            #else:  #debug code
                #cellVal=cell.value.lower()
                #print(cellVal)
                #srcLang=targetLanguageRaw.lower()
                #print(srcLang)
                #print('"'+cellVal+'"!="'+srcLang+'"')

if targetLanguageCell == None:
    sys.exit(('\n Error. The following language was not found in"' + languageCodesFileName + '":"' + targetLanguageRaw+'"').encode(consoleEncoding))
else:
    print(('TargetLanguage:'+targetLanguageRaw+':'+str(targetLanguageCell)).encode(consoleEncoding))
#Assume targetLanguageCell has a valid value now.
targetLanguageRow, targetLanguageColumn = getRowAndColumnFromCell(targetLanguageCell)
#print(languageCodesSpreadsheet['A'+targetLanguageRow].value)

if debug == True:
    print((str(languageCodesSpreadsheet['A'+targetLanguageRow].value)).encode(consoleEncoding))
    print((str(languageCodesSpreadsheet['B'+targetLanguageRow].value)).encode(consoleEncoding))
    print((str(languageCodesSpreadsheet['C'+targetLanguageRow].value)).encode(consoleEncoding))
    print((str(languageCodesSpreadsheet['D'+targetLanguageRow].value)).encode(consoleEncoding))
    print((str(languageCodesSpreadsheet['E'+targetLanguageRow].value)).encode(consoleEncoding))
    print((str(languageCodesSpreadsheet['F'+targetLanguageRow].value)).encode(consoleEncoding))
    print((str(languageCodesSpreadsheet['G'+targetLanguageRow].value)).encode(consoleEncoding))
    print((str(languageCodesSpreadsheet['H'+targetLanguageRow].value)).encode(consoleEncoding))

#if str(languageCodesSpreadsheet['E'+targetLanguageRow].value) == 'False':

#The destination language does not have its source language code modified, so just set the values unconditionally.
internalLanguageDestinationName=languageCodesSpreadsheet['A'+targetLanguageRow].value
internalLanguageDestinationTwoCode=languageCodesSpreadsheet['B'+targetLanguageRow].value
internalLanguageDestinationThreeCode=languageCodesSpreadsheet['C'+targetLanguageRow].value

#elif str(languageCodesSpreadsheet['E'+targetLanguageRow].value) == 'True':
#    internalLanguageDestinationName=languageCodesSpreadsheet['F'+targetLanguageRow].value
#    internalLanguageDestinationTwoCode=languageCodesSpreadsheet['G'+targetLanguageRow].value
#    internalLanguageDestinationThreeCode=languageCodesSpreadsheet['H'+targetLanguageRow].value
#else:
#    print('')
#    print(str(languageCodesSpreadsheet['E'+targetLanguageRow].value).encode(consoleEncoding))
#    sys.exit('Unspecified error.'.encode(consoleEncoding))

#if debug == True:
#    print((str(bool(languageCodesSpreadsheet['E'+sourceLanguageRow].value))).encode(consoleEncoding))
#    print(str(internalLanguageDestinationName).encode(consoleEncoding))
#    print(str(internalLanguageDestinationTwoCode).encode(consoleEncoding))
#    print(str(internalLanguageDestinationThreeCode).encode(consoleEncoding))
"""

#Read in all dictionaries.
#read in character dictionary


#read in pre dictionary


#read in post dictionary


#read in afterWritingToFile dictionary










#debug code
#print(inputFileHandle.read().encode(consoleEncoding))
#inputFileHandle.close()  #tidy up


#CharaNamesDictionary time to shine
CharacterNames=['[＠クロエ]','Chloe']#change this to a dictionary
#If an ignored line starts with a character name, that could create problems, so swap names first, then parse lines into main database.
#if one was specified, import character definitions file, and perform swaps while inputfile is still in memory but has not yet been parsed line by line.
#inputFileContents.replaceStuffHere()

#somehow this line needs to run
#inputFileContents=inputFileContents.replace('[＠クロエ]','Chloe')


#Import parse settings from parsingDefintionsFile, if parsing is required which is if....
#if... parseOnly mode specified, resume has not specified (resume implies there is a file under backup/)
#always parse regardless of mode if one is specified, and then import data from other sources as it becomes available. Keep track of whether or not main data structure exists, or is supposed to be created from resume data (backup/ folder, csv, xls/xlsx, ods).
ignoreLinesThatStartWith=[ '[' , '*' , ';' , '【' ]     #this is a list of strings where each entry contains a character that marks a line to skip processing it, only one character is considered per entry
paragraphDelimiter='emptyLine'         #Valid options: emptyLine, newLine  #Once input has started, when should it stop? Input will stop automatically regardless of this setting when maximumParagraphSize is reached.
maximumNumberOfLinesPerParagraph=3        #Maximum number of lines per paragraph. A sane setting for this is 2-5 lines maximum per paragraph.
#What is the maximum length for translated text before triggering word wrap? The correct setting for this depends on
#the font used, which can sometimes be changed by the user at runtime, and the target language.
#Sane values are 30-60. This setting strongly influences the number of calculated output lines.
wordWrap=45
#The number of untranslated input lines will not always match the number of translated output lines, especially with
#different wordWrap amounts. How should this situation be handled? Valid values are: none, strict, dynamic.
#none=Disable word wrapping and always dump all translated text onto one line. Replace subsequent untranslated lines with empty lines.
#strict=If the number of input lines and translated lines match exactly, then replace as normal. Otherwise, do nothing.
    #The lines with a mismatch will be placed into a 'mixmatch.xlsx' file, and it is the user's responsibility to sort through those lines.
#dynamic=If there are fewer lines after translation, replace the extra untranslated lines with empty lines.
    #If there are more lines after translation, then append the extra lines to the last untranslated line.
wordWrapMode='dynamic'


#initialize main data structure
#mainDatabaseWorkbook = Workbook()
#mainDatabaseSpreadsheet = mainDatabaseWorkbook.active

#Strawberry is a wrapper class for the workbook class with additional methods.
#The interface has no concept of workbooks vs spreadsheets. That distinction is handled only inside the class.
mainSpreadsheet=Strawberry()


#This will hold raw untranslated text, number of lines each entry was derived from, and columns containing maybe translated data. Maximum columns supported = 9[?]. Columns 0 and 1 are reserved for Raw untranslated text and [1] is reserved for the number of lines that column 0 represents, even if it is always 1 because line-by-line parsing (lineByLineMode) is specified.
#Alternatively, lineByLineMode can instead specify to feed each line into the translator but the paragraphDelimiter=emptyLine can support multiple lines in the main data structure. That might be more flexible, or make things worse. Leave it to user to decide on a case-by-case basis by allowing for this possibility.
#class Database:
#    def __init__(self):
#This data structure must somehow hold:
#0) each paragraph.raw.text (with \n's inside of it)
#1) The speaker of the line, if any.
#2) metadata (# of lines #1 was taken from, and possibly the line number of the entry, possibly multiple entries metadata, possibly arbitary data to be decided later like booleans or whether the line has been translated or not), maybe a list?
#Translated or not decision should be computed dynamically and on a line-by-line basis (raw + chosen translator engine determined from header) in order to support resume operations, do not add to metadata
#For now use the following string:  'numberOfSourceLines!WasADictionaryUsedTrueOrFalse' ex. '2!False!'
#3) all translations, each with its own entry and an associated header for which type it is (Kobold+model;
#for DeepL, DeepL API Free/Pro and DeepL Web should all be the same right? Free, Pro, and Web versions all support dictionaries, but doing dictionary + document translation requires pro. Documents will be parsed prior to using the API, so this limitation does not apply here, thus all DeepL translations should be the same. DeepL's native dictionary system is also unlikely to be implemented here especially considering they are hard to work with (not mutable).
#4) potentially, koboldcpp LLM could return different models. Use the returned model header name. All data should be put in that column. Determine correct header as in: DeepL, or koboldcpp/model in a case sensitive way. More examples:
#headers=['rawText', 'metadata', 'DeepL', 'koboldcpp/Mixtral-13B.Q8_0','Google']

#Add headers
initialHeaders=['rawText','metadata']
mainDatabaseSpreadsheet.append(initialHeaders)

#printAllTheThings(mainDatabaseSpreadsheet)

#start parsing input file line by line
#print(inputFileContents.encode(consoleEncoding))
#print(inputFileContents.partition('\n')[0].encode(consoleEncoding)) #prints only first line



#parseRawInputTextFile() accepts an (input file name, the encoding for that text file, parseFileDictionary as a Python dictionary, the character dictionary as a Python dictionary) and returns a dictionary where the key is the dialogue, and the value is a list. The first value in the list is the character name, (or None for no chara name), and the second is metadata as a string using the specified delimiter.
#This could also be a multidimension array, such as a list full of list pairs [ [ [],[ [][][] ] ] , [ [],[ [][][] ] ] ]  because the output is highly regular, but that would allow duplicates.  Executive decision was made to disallow duplicates for files since that is correct almost always. However, it does mess with the metadata sometimes by having the speaker be potentially incorrect.
temporaryDict=parseRawInputTextFile(inputFile,inputFileEncoding,parseFileDictionary, characterDictionary)

temporaryDict={}        #Dictionaries do not allow duplicates, so insert all entries into a dictionary first to de-duplicate entries, then read dictionary into first column (skip first line/row in target spreadsheet)
#thisdict.update({"x": "y"}) #add to/update dictionary
#thisdict["x"]="y"              #add to/update dictionary
#for x, y in thisdict.items():
#  print(x, y) 


currentRow=2
#for ever item in the dictionary
#print(str(temporaryDict).encode(consoleEncoding))
for text, metadata in temporaryDict.items():
    #print(text.encode(consoleEncoding))
    #print(metadata.encode(consoleEncoding))
    #add item to spreadsheet column1 (A) and incrementing rows starting with row #2
    mainDatabaseSpreadsheet['A'+str(currentRow)]=text
    mainDatabaseSpreadsheet['B'+str(currentRow)]=metadata
    #then move on to next item in dictionary and next row
    currentRow+=1

if debug == True:
    printAllTheThings(mainDatabaseSpreadsheet)

#create backup of database when it is initially created
print('')
#print(('creating backup at: backups/'+currentDateFull+'/rawUntranslated-'+currentDateAndTimeFull+'.xlsx').encode(consoleEncoding))
print('')
Path('backups/'+currentDateFull).mkdir(parents=True, exist_ok=True)
#mainDatabaseWorkbook.save('backups/'+currentDateFull+'/rawUntranslated-'+currentDateAndTimeFull+'.xlsx')
#print(inputFileContents.partition('\n')[0].encode(consoleEncoding)) #prints only first line






#once all lines are input into a dictionary, if specified, read the preDictionary.csv and perform replacements prior to submission to translation engine


#Then run DeepL's special code to replace braces {{ }}  {{{ }}}  with special markers that the DeepL engine can escape before submission to the translation engine


#




#do in memory replacements for CharaNames
#Replace [＠クロエ] with Chloe, but keep a record that it was replaced in order to replace it back after translation.
#Should be read from a c.csv that includes
#1) Information that should be preserved in the final translation as-is, but that will also be submitted to the translation engine for consideration so the position of the name or other string can influence the translation itself.
#To alter specific strings in the text before submission to the translation engine that will also be kept in the translated text, use preDictionary.csv. To alter specific parts of the text after translation, use postDictionary.csv.

#For character names, either:
#need to know, gender (m, f, u), since that might influence the translation.  Can also append that information to prompt for LLM models.
#Could also ask user to specify a replacement name. This might be a better idea.



#The input file must be either a raw file (.ks, .txt, .ts) or a spreadsheet file (.csv, .xlsx, .xls, .ods), so the encoding for the spreadsheets... is that always utf-8?
#Raw files can be spreadsheets with one column, but spreadsheet files cannot be raw files. So include a switch that signifies to process text files as psudo .csv spreadsheets? No. Just have user convert them to .csv manually instead. Then no special handling is needed for 'special' \n deliminated .csv's. This feature can be added later if requested.
#However, when using files exported from other programs. There will be no metadata. Instead of adding a cli flag, fallback to line-by-line mode instead and warn user about it. Fallback to line by line mode should also occur if the metadata is corrupted. Is that really necessary? Lines count can be found based upon column 1 entries, and word wrap settings only need the parsedefintions file. That is sperate from translating, which can still be done using previously obtained paragraphs. Speaker will be missing.




ignoreLinesThatStartWith=['[','*',';','【','・']
myDict={'[＠クロエ]':'Chloe'}

print('myDict='+str(myDict))
longString=' '+'[＠クロエ]'+',dawd j3 a!@$%(^&  '
shortString=' [＠ab '
print(longString)
print(shortString+'\n')

#this will print the full string without returning an error
x = 'pie2'
print(x[:9])

thisLineIsValid=True
#How to check if dictionaryEntry is at the start of both short and longString?
for i in ignoreLinesThatStartWith:
    if longString.strip()[:1] == i:
        print('matching Value: \''+i+'\'')
        thisLineIsValid=False
        for j,k in myDict.items():
            if j[:1] == i:
               #Then it might match if the first character matches.
               #print('pie')
               #if the dictionary entry is the same as the start of the line
               print('dictionaryKeyEntry='+j)
               print( 'rawLine=\''+longString+'\'')
               print( 'cleanedUpLine='+longString.strip()[:len(j)] )
               if j == longString.strip()[:len(j)]:
                   print('pie3')
                   thisLineIsValid=True

print('\nthisLineIsValid='+str(thisLineIsValid))


